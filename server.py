"""
PFC LOGISTICS ROUTING SYSTEM - Server v3
Premier Food Choice | Automated Delivery Routing
Changes v3:
  - Cluster-first routing engine (same cluster → same truck, avoids last-mile fees)
  - Truck merging pass (reduce truck count, maximise utilisation)
  - TSP nearest-neighbor stop sequencing (minimise drop distance)
  - Rate Master DB table + CRUD API
  - Monitoring records DB persistence + analytics
  - Per-router temp files (multi-user support)
  - Cost/KG per truck via Rate Master lookup
"""

from flask import Flask, request, jsonify, send_file, render_template, session
import sqlite3
import json
import math
import re
import os
import io
import tempfile
import urllib.request
import urllib.parse
from datetime import datetime
from collections import defaultdict
from werkzeug.utils import secure_filename

# Lazy pandas import — avoids crashing the whole server if pip install was incomplete
def _get_pd():
    try:
        import pandas as _pd
        return _pd
    except ImportError:
        raise RuntimeError(
            "pandas is not installed. Please run: pip install pandas openpyxl xlrd"
        )

app = Flask(__name__)

import hashlib as _hashlib
# Secret key: reads from environment on hosted server, falls back to dev default
import os as _os
app.secret_key = _os.environ.get('SECRET_KEY', 'pfc-logistics-secret-key-2026-x9z')

def _hw(pw):
    return _hashlib.sha256(pw.encode('utf-8')).hexdigest()

_USERS = {
    'planner1': {'pw': _hw('PFCplan1!'), 'role': 'planner', 'display': 'Route Planner 1'},
    'planner2': {'pw': _hw('PFCplan2!'), 'role': 'planner', 'display': 'Route Planner 2'},
    'planner3': {'pw': _hw('PFCplan3!'), 'role': 'planner', 'display': 'Route Planner 3'},
    'tower1':   {'pw': _hw('PFCtower1!'), 'role': 'tower',   'display': 'Logistics Tower 1'},
    'tower2':   {'pw': _hw('PFCtower2!'), 'role': 'tower',   'display': 'Logistics Tower 2'},
    'tower3':   {'pw': _hw('PFCtower3!'), 'role': 'tower',   'display': 'Logistics Tower 3'},
}

_AUTH_EXEMPT = {'/api/login', '/api/logout', '/api/me', '/api/warehouses', '/api/rates'}
_TOWER_BLOCKED = {'/api/upload', '/api/run-routing', '/api/merge-orders',
                  '/api/save-plan', '/api/delete-plan'}

@app.before_request
def _check_auth():
    if request.path in _AUTH_EXEMPT:
        return None
    if request.path == '/' or not request.path.startswith('/api/'):
        return None
    if not session.get('username'):
        return jsonify({'error': 'Not authenticated', 'login_required': True}), 401
    if session.get('role') == 'tower' and request.path in _TOWER_BLOCKED:
        return jsonify({'error': 'Access denied for your role'}), 403
    return None

@app.route('/api/login', methods=['POST'])
def api_login():
    data = request.get_json() or {}
    username = (data.get('username') or '').strip().lower()
    password = (data.get('password') or '').strip()
    user = _USERS.get(username)
    if not user or user['pw'] != _hw(password):
        return jsonify({'error': 'Invalid username or password'}), 401
    session.permanent = True
    session['username'] = username
    session['role']     = user['role']
    session['display']  = user['display']
    return jsonify({'ok': True, 'username': username,
                    'role': user['role'], 'display': user['display']})

@app.route('/api/logout', methods=['POST'])
def api_logout():
    session.clear()
    return jsonify({'ok': True})

@app.route('/api/me', methods=['GET'])
def api_me():
    if not session.get('username'):
        return jsonify({'authenticated': False}), 401
    return jsonify({'authenticated': True,
                    'username': session['username'],
                    'role':     session['role'],
                    'display':  session['display']})

_collab = {}  # collab_id -> {'trucks':[...], 'version':float, 'by':str}
_DATA_DIR = os.environ.get('DATA_DIR', os.path.dirname(os.path.abspath(__file__)))
app.config['UPLOAD_FOLDER'] = os.path.join(_DATA_DIR, 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024

DB_PATH = os.path.join(_DATA_DIR, 'pfc_logistics.db')
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

def open_db():
    """Open SQLite with WAL mode + generous busy timeout to prevent 'database is locked' errors."""
    conn = sqlite3.connect(DB_PATH, timeout=30, check_same_thread=False)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=30000")   # 30 s retry window
    conn.execute("PRAGMA synchronous=NORMAL")   # safe + faster than FULL
    return conn

# ─────────────────────────────────────────────────────────────
#  WAREHOUSE DATA
# ─────────────────────────────────────────────────────────────
# Warehouse groups that are NOT auto-routed — stops are floated to unrouted basket
FLOAT_WH_GROUPS = {'FLOAT'}

WAREHOUSE_DATA = {
    'FCSC': {'lat': 14.638413, 'lng': 120.955994, 'name': 'Frabelle Cold Storage (Navotas)',  'group': 'NAVOTAS'},
    'CCS1': {'lat': 14.640139, 'lng': 120.955142, 'name': 'Crystal Cold Storage 1 (Navotas)', 'group': 'NAVOTAS'},
    'CCS2': {'lat': 14.640500, 'lng': 120.955200, 'name': 'Crystal Cold Storage 2 (Navotas)', 'group': 'NAVOTAS'},
    'MLIB': {'lat': 14.841508, 'lng': 120.903061, 'name': 'Mets Cold Storage - Bulacan',      'group': 'BULACAN'},
    'MLIC': {'lat': 14.289947, 'lng': 121.014182, 'name': 'Mets Cold Storage - Cavite',       'group': 'CARMONA'},
    'SACS': {'lat': 14.310108, 'lng': 121.035113, 'name': 'South Alps Cold Storage (Carmona)','group': 'CARMONA'},
    'FGCS': {'lat': 14.638413, 'lng': 120.955994, 'name': 'FGCS (Frabelle Navotas)',           'group': 'NAVOTAS'},
    # Float warehouses — stops are NOT auto-routed; planner assigns manually
    'RGM':  {'lat': 14.5995,  'lng': 120.9842,   'name': 'RGM Head Office',                    'group': 'FLOAT'},
}

# WH group → Rate Master "Pickup_WH" column value
WH_TO_RATE_WH = {
    'NAVOTAS': 'NAVOTAS',
    'BULACAN': 'BALAGTAS',
    'CARMONA': 'CARMONA',
    'FGCS':    'NAVOTAS',
}

def get_wh_group(wh_code: str) -> str:
    return WAREHOUSE_DATA.get(wh_code, {}).get('group', wh_code)

def get_wh_coords(wh_code: str):
    d = WAREHOUSE_DATA.get(wh_code, {})
    return d.get('lat', 14.55), d.get('lng', 121.00)

# ─────────────────────────────────────────────────────────────
#  TRUCK TYPES
# ─────────────────────────────────────────────────────────────
TRUCK_TYPES = [
    {'label': '2T',    'cap': 2000,  'max_drops': 8},
    {'label': '2.5MT', 'cap': 2500,  'max_drops': 10},
    {'label': '5MT',   'cap': 5000,  'max_drops': 10},
    {'label': '10MT',  'cap': 10000, 'max_drops': 12},
    {'label': '15MT',  'cap': 15000, 'max_drops': 15},
    {'label': '20MT',  'cap': 20000, 'max_drops': 18},
    {'label': '27MT',  'cap': 27000, 'max_drops': 20},
]

SETTINGS = {
    'util_target':               0.80,
    'cap_tolerance':             1.05,
    'two_ton_max_util':          1.20,   # 2T trucks allow up to 120% (2400 kg)
    'add_drop_fee':              300,
    'add_drop_threshold_5mt':    5,
    'add_drop_threshold_25mt':   7,
    'prefer_5mt_min_vol':        2000,
}

def select_truck_type(vol: float) -> dict:
    """Smallest truck type that fits the given volume.
    Routes ≤ 2000 kg always get 2T.
    """
    if vol > 20000: return TRUCK_TYPES[5]
    if vol > 15000: return TRUCK_TYPES[4]
    if vol > 10000: return TRUCK_TYPES[3]
    if vol > 5000:  return TRUCK_TYPES[2]
    if vol > TRUCK_TYPES[0]['cap']: return TRUCK_TYPES[1]  # 2000–5000 kg → 2.5MT
    return TRUCK_TYPES[0]  # ≤ 2000 kg → 2T

def best_truck_type(vol: float, drops: int) -> dict:
    """Smallest truck type that fits both volume AND drop count.
    2T trucks are preferred for routes ≤2400 kg (120% utilization allowed).
    If actual volume is under 2000 kg, always assign 2T regardless of drop count.
    """
    two_ton_max = TRUCK_TYPES[0]['cap'] * SETTINGS['two_ton_max_util']  # 2400 kg
    if vol <= TRUCK_TYPES[0]['cap']:  # vol ≤ 2000 kg → always 2T
        return TRUCK_TYPES[0]
    for t in TRUCK_TYPES:
        tol = SETTINGS['two_ton_max_util'] if t['label'] == '2T' else SETTINGS['cap_tolerance']
        if t['cap'] * tol >= vol and t['max_drops'] >= drops:
            return t
    return TRUCK_TYPES[-1]

# ─────────────────────────────────────────────────────────────
#  COLUMN NORMALISER
# ─────────────────────────────────────────────────────────────
def normalize_headers(columns: list) -> dict:
    mapping = {}
    for col in columns:
        lc = str(col).lower().strip()
        if lc in ('customer', 'customer name') or lc.startswith('customer name') or \
           (lc.startswith('customer') and 'name' in lc):
            mapping.setdefault('customer_name', col)
        elif 'ship to address' in lc or 'shipping address' in lc:
            mapping.setdefault('shipping_address', col)
        elif 'so tfor qty' in lc or lc == 'tfor qty':
            mapping.setdefault('tfor_qty', col)
        elif lc == 'do qty' or 'do qty' in lc or \
             ('qty' in lc and 'kilo' in lc) or lc == 'do qty (in kilos)':
            mapping.setdefault('do_qty', col)
        elif lc in ('location', 'pickup wh'):
            mapping.setdefault('location', col)
        elif 'cluster id' in lc or lc == 'so cluster id':
            mapping.setdefault('cluster_id', col)
        elif 'route id' in lc or lc == 'so route id':
            mapping.setdefault('route_id', col)
        elif lc == 'so delivery area':
            mapping['area'] = col
        elif lc == 'area' and 'area' not in mapping:
            mapping['area'] = col
        elif 'latitude' in lc or lc in ('lat', 'so latitude', 'delivery lat', 'delivery latitude'):
            mapping.setdefault('delivery_latitude', col)
        elif 'longitude' in lc or 'longtitude' in lc or \
             lc in ('lng', 'lon', 'so longitude', 'delivery lng', 'delivery longitude', 'delivery longtitude'):
            mapping.setdefault('delivery_longitude', col)
        elif lc in ('so number', 'document number', 'so_number'):
            mapping.setdefault('so_number', col)
        elif lc == 'ship date':
            mapping.setdefault('ship_date', col)
        elif 'blanket' in lc:
            mapping.setdefault('blanket_so_date', col)
        elif 'shipping method' in lc:
            mapping.setdefault('shipping_method', col)
        elif 'date and time' in lc or 'date encoded' in lc:
            mapping.setdefault('date_encoded', col)
        elif 'internal id' in lc:
            mapping.setdefault('internal_id', col)
    return mapping

# ─────────────────────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────────────────────
def clean_area_name(text: str) -> str:
    """Return the primary area token from a cluster code like NCR-RIZ-TYT → RIZ."""
    if not text: return ''
    clean = str(text).upper().strip()
    # Remove NCR- prefix first
    clean = re.sub(r'^NCR\s*[-–]\s*', '', clean).strip()
    # Remove trailing "CITY"
    clean = re.sub(r'\s*CITY\s*$', '', clean, flags=re.IGNORECASE).strip()
    # If still hyphenated (e.g. RIZ-TYT, CAL-SW), take the FIRST segment (province/city code)
    if '-' in clean:
        clean = clean.split('-')[0].strip()
    return clean


def area_candidates(text: str) -> list:
    """Return all area name candidates to try in rate master (primary first)."""
    if not text: return []
    raw = str(text).upper().strip()
    seen, result = set(), []
    def add(s):
        s = s.strip()
        if s and s not in seen:
            seen.add(s); result.append(s)
    # Primary cleaned form
    add(clean_area_name(raw))
    # Without NCR prefix, full remainder
    no_ncr = re.sub(r'^NCR\s*[-–]\s*', '', raw).strip()
    no_ncr = re.sub(r'\s*CITY\s*$', '', no_ncr, flags=re.IGNORECASE).strip()
    add(no_ncr)
    # All hyphen segments
    for part in no_ncr.split('-'):
        add(part.strip())
    # Original raw (stripped)
    add(raw)
    return result


def parse_location_cell(loc_raw: str) -> str:
    """Extract the pickup WH code from a Location cell.

    The cell format can be:
      - 'PSC : MLIC Good'   → pickup WH = MLIC  (first part is customer CSW label)
      - 'MLIB : MLIB Good'  → pickup WH = MLIB
      - 'FCSC'              → pickup WH = FCSC   (plain code, no colon)
      - 'MLIC Good'         → pickup WH = MLIC   (no colon, WH + status)

    Logic: scan all whitespace-delimited tokens in the ENTIRE cell and return
    the first one that matches a known WAREHOUSE_DATA key.  If none match,
    fall back to the token after the colon (if present), else the first token.
    """
    raw = str(loc_raw).strip().upper()
    # Tokenise: split on colon and spaces, keep all words
    tokens = [t.strip().rstrip('.') for t in re.split(r'[\s:]+', raw) if t.strip()]

    # Priority 1: any token that is a known WH key
    for tok in tokens:
        if tok in WAREHOUSE_DATA:
            return tok

    # Priority 2: token after the colon (second logical segment)
    if ':' in raw:
        after = raw.split(':', 1)[1].strip()
        first_after = after.split()[0].rstrip('.') if after.split() else ''
        if first_after:
            return first_after

    # Priority 3: first token
    return tokens[0] if tokens else 'UNKNOWN'


def haversine_distance(lat1, lon1, lat2, lon2) -> float:
    try:
        lat1, lon1, lat2, lon2 = float(lat1), float(lon1), float(lat2), float(lon2)
        if not (lat1 and lon1 and lat2 and lon2): return 9999
        R = 6371
        dlat = math.radians(lat2 - lat1)
        dlon = math.radians(lon2 - lon1)
        a = (math.sin(dlat / 2) ** 2
             + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2))
             * math.sin(dlon / 2) ** 2)
        return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    except Exception:
        return 9999

def safe_get(row: dict, col_map: dict, key: str, default=''):
    col = col_map.get(key)
    if col and col in row:
        val = row[col]
        if val is None or (isinstance(val, float) and math.isnan(val)):
            return default
        return val
    return default

def serialise_row(row_dict: dict) -> dict:
    out = {}
    for k, v in row_dict.items():
        if isinstance(v, float) and math.isnan(v): out[k] = ''
        elif hasattr(v, 'isoformat'): out[k] = str(v)
        else: out[k] = v
    return out

def get_router_temp(suffix: str, router: str = 'default') -> str:
    """Return path to router-specific temp JSON file."""
    safe = re.sub(r'[^a-zA-Z0-9_-]', '_', str(router).lower().strip() or 'default')
    return os.path.join(tempfile.gettempdir(), f'pfc_{suffix}_{safe}.json')


# ── Shared live working plan helpers ────────────────────────────────────────
def get_working_plan(plan_date: str):
    """Read the shared live plan for a date from the DB.  Returns (trucks, updated_by, updated_at)."""
    try:
        conn = open_db(); c = conn.cursor()
        c.execute('SELECT data, updated_by, updated_at FROM working_plans WHERE plan_date=?', (plan_date,))
        row = c.fetchone(); conn.close()
        if not row: return None, '', ''
        return json.loads(row[0]), row[1] or '', row[2] or ''
    except Exception:
        return None, '', ''


def set_working_plan(plan_date: str, trucks: list, updated_by: str = ''):
    """Write / update the shared live plan for a date in the DB."""
    try:
        conn = open_db(); c = conn.cursor()
        c.execute('''INSERT INTO working_plans (plan_date, data, updated_by, updated_at)
                     VALUES (?, ?, ?, datetime('now','localtime'))
                     ON CONFLICT(plan_date) DO UPDATE SET
                         data       = excluded.data,
                         updated_by = excluded.updated_by,
                         updated_at = datetime('now','localtime')''',
                  (plan_date, json.dumps(trucks, default=str), updated_by))
        conn.commit(); conn.close()
    except Exception as ex:
        print(f'[WARN] set_working_plan failed: {ex}')
# ────────────────────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────
#  TSP: NEAREST-NEIGHBOR STOP SEQUENCING
# ─────────────────────────────────────────────────────────────
def nearest_neighbor_tsp(start_lat: float, start_lng: float, stops: list) -> list:
    """Reorder stops starting from warehouse using nearest-neighbor heuristic."""
    if len(stops) <= 1:
        return stops
    remaining = stops[:]
    result = []
    cur_lat, cur_lng = start_lat, start_lng
    while remaining:
        best = min(remaining,
                   key=lambda s: haversine_distance(cur_lat, cur_lng,
                                                    s.get('lat', 0), s.get('lng', 0)))
        result.append(best)
        cur_lat = best.get('lat', cur_lat)
        cur_lng = best.get('lng', cur_lng)
        remaining.remove(best)
    return result

# ─────────────────────────────────────────────────────────────
#  RATE MASTER LOOKUP
# ─────────────────────────────────────────────────────────────
def lookup_rate(trucker: str, capacity_kg: int, wh_group: str, area: str):
    """Returns (rate, cost_per_kg) from DB or (None, None) if not found."""
    if not trucker:
        return None, None
    rate_wh     = WH_TO_RATE_WH.get(wh_group, wh_group).upper()
    area_clean  = clean_area_name(area).upper()
    trucker_up  = trucker.upper().strip()
    cap         = int(capacity_kg)

    try:
        conn = open_db()
        c = conn.cursor()
        # Exact match: trucker + capacity + wh + area
        c.execute('''SELECT rate, cost_per_kg FROM rate_master
                     WHERE UPPER(trucker)=? AND capacity_kg=?
                       AND UPPER(pickup_wh)=? AND UPPER(area)=?
                       AND is_active=1 LIMIT 1''',
                  (trucker_up, cap, rate_wh, area_clean))
        row = c.fetchone()
        if not row:
            # Fallback: trucker + capacity + wh only (any area)
            c.execute('''SELECT rate, cost_per_kg FROM rate_master
                         WHERE UPPER(trucker)=? AND capacity_kg=?
                           AND UPPER(pickup_wh)=? AND is_active=1
                         ORDER BY id LIMIT 1''',
                      (trucker_up, cap, rate_wh))
            row = c.fetchone()
        conn.close()
        if row:
            return float(row[0] or 0), float(row[1] or 0)
    except Exception:
        pass
    return None, None

def calculate_total_km(wh_lat: float, wh_lng: float, stops: list) -> float:
    """Sum haversine distances: WH → stop1 → stop2 → ... → last stop."""
    total = 0.0
    cur_lat, cur_lng = wh_lat, wh_lng
    for s in stops:
        slat = float(s.get('lat', 0) or 0)
        slng = float(s.get('lng', 0) or 0)
        if slat and slng:
            total += haversine_distance(cur_lat, cur_lng, slat, slng)
            cur_lat, cur_lng = slat, slng
    return round(total, 2)


def lookup_rate_for_area(c, trucker_up, cap, rate_wh, area_raw):
    """Try all candidate area names for one area string. Returns (rate, cpk) or (None, None)."""
    for candidate in area_candidates(str(area_raw)):
        c.execute('''SELECT rate, cost_per_kg FROM rate_master
                     WHERE UPPER(trucker)=? AND capacity_kg=?
                       AND UPPER(pickup_wh)=? AND UPPER(area)=?
                       AND is_active=1 LIMIT 1''',
                  (trucker_up, cap, rate_wh, candidate.upper()))
        row = c.fetchone()
        if row:
            return float(row[0] or 0), float(row[1] or 0)
    return None, None


def lookup_highest_rate(trucker: str, capacity_kg: int, wh_group: str, areas: list):
    """Return (max_rate, cost_per_kg) for the highest-rate area served by this truck."""
    if not trucker:
        return None, None
    rate_wh    = WH_TO_RATE_WH.get(wh_group, wh_group).upper()
    trucker_up = trucker.upper().strip()
    cap        = int(capacity_kg)
    try:
        conn = open_db()
        c    = conn.cursor()
        best_rate = None
        best_cpk  = None
        for area in set(areas):
            r, cpk = lookup_rate_for_area(c, trucker_up, cap, rate_wh, area)
            if r is not None and (best_rate is None or r > best_rate):
                best_rate = r
                best_cpk  = cpk
        if best_rate is None:
            # Fallback 1: any area for this trucker+capacity+wh
            c.execute('''SELECT rate, cost_per_kg FROM rate_master
                         WHERE UPPER(trucker)=? AND capacity_kg=?
                           AND UPPER(pickup_wh)=? AND is_active=1
                         ORDER BY rate DESC LIMIT 1''',
                      (trucker_up, cap, rate_wh))
            row = c.fetchone()
            if row:
                best_rate = float(row[0] or 0)
                best_cpk  = float(row[1] or 0)
        if best_rate is None and cap > 0:
            # Fallback 2: ignore capacity — any rate for this trucker+wh
            # (handles case where rate master uses different cap values)
            c.execute('''SELECT rate, cost_per_kg FROM rate_master
                         WHERE UPPER(trucker)=? AND UPPER(pickup_wh)=? AND is_active=1
                         ORDER BY rate DESC LIMIT 1''',
                      (trucker_up, rate_wh))
            row = c.fetchone()
            if row:
                best_rate = float(row[0] or 0)
                best_cpk  = float(row[1] or 0)
        if best_rate is None:
            # Fallback 3: any rate for this trucker (ignore wh and capacity)
            c.execute('''SELECT rate, cost_per_kg FROM rate_master
                         WHERE UPPER(trucker)=? AND is_active=1
                         ORDER BY rate DESC LIMIT 1''',
                      (trucker_up,))
            row = c.fetchone()
            if row:
                best_rate = float(row[0] or 0)
                best_cpk  = float(row[1] or 0)
        conn.close()
        if best_rate is not None:
            return best_rate, best_cpk
    except Exception:
        pass
    return None, None


# ─────────────────────────────────────────────────────────────
#  GEOCODING HELPERS  (used by routing engine + fix-coords)
# ─────────────────────────────────────────────────────────────
def _address_candidates(addr: str) -> list:
    """Return a list of progressively-simplified address strings to try with Nominatim."""
    import re as _re
    raw = addr.strip()
    candidates = []
    seen = set()

    def add(s):
        s = s.strip()
        if s and s not in seen:
            seen.add(s)
            candidates.append(s)

    # 1. Full address as-is
    add(raw)
    # 2. Drop leading unit/floor numbers (e.g. "Unit 3A, ..." → "...")
    no_unit = _re.sub(r'^(unit|floor|flr|bldg|blk|lot|rm|room)[\s\d\w]*[,\s]+',
                      '', raw, flags=_re.IGNORECASE).strip()
    add(no_unit)
    # 3. Last two comma-separated parts (street + city)
    parts = [p.strip() for p in raw.split(',') if p.strip()]
    if len(parts) >= 2:
        add(', '.join(parts[-2:]))
    # 4. Last part only (city/municipality)
    if parts:
        add(parts[-1])

    return candidates


def _nominatim_search(query: str) -> dict | None:
    """Send one search request to Nominatim; return the best hit dict or None."""
    import urllib.request as _ureq, urllib.parse as _uparse, json as _json
    params = _uparse.urlencode({
        'q':              query,
        'format':         'json',
        'limit':          1,
        'countrycodes':   'ph',
        'addressdetails': 0,
    })
    url = f'https://nominatim.openstreetmap.org/search?{params}'
    req = _ureq.Request(url, headers={
        'User-Agent': 'PFC-Routing/1.0 supplychain3@premierfoodchoice.com'
    })
    try:
        with _ureq.urlopen(req, timeout=8) as resp:
            data = _json.loads(resp.read())
        return data[0] if data else None
    except Exception:
        return None


# ─────────────────────────────────────────────────────────────
#  ROUTING ENGINE  (v4: proximity-first + TSP + total_km)
# ─────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────
#  CLARKE-WRIGHT SAVINGS ALGORITHM
# ─────────────────────────────────────────────────────────────
def clarke_wright_routing(wh_lat, wh_lng, stops, cap_tolerance=None, target_cap=5000):
    """
    Clarke-Wright Savings Algorithm for Capacitated VRP.

    For every pair of stops (i, j) computes the saving achieved by
    serving them on the same route instead of two separate runs:
        s(i,j) = d(WH,i) + d(WH,j) - d(i,j)

    Pairs are merged greedily (highest saving first) subject to:
      - Combined volume <= target_cap * cap_tolerance (default: 5MT)
      - Combined drop count <= 10
      - Both stops must be at the *ends* of their current routes

    Stops whose individual volume already exceeds target_cap are passed in
    as single-stop routes so best_truck_type() can assign them a larger truck.

    Returns a list of routes; each route is a list of stop dicts.
    The caller is responsible for TSP-reordering each route.
    """
    if cap_tolerance is None:
        cap_tolerance = SETTINGS['cap_tolerance']

    if not stops:
        return []
    if len(stops) == 1:
        return [list(stops)]

    n = len(stops)

    # Safe coord accessors (fall back to WH if stop has no coords)
    def slat(s): return s['lat']  if s.get('lat') else wh_lat
    def slng(s): return s['lng']  if s.get('lng') else wh_lng

    # Distance from WH to each stop (pre-computed)
    d_wh = [haversine_distance(wh_lat, wh_lng, slat(stops[i]), slng(stops[i]))
            for i in range(n)]

    # Build savings list: (saving_value, i, j)
    savings = []
    for i in range(n):
        for j in range(i + 1, n):
            d_ij = haversine_distance(slat(stops[i]), slng(stops[i]),
                                      slat(stops[j]), slng(stops[j]))
            s = d_wh[i] + d_wh[j] - d_ij
            savings.append((s, i, j))
    savings.sort(key=lambda x: -x[0])   # highest saving first

    # Initialise: each stop is its own single-stop route
    routes    = [[i] for i in range(n)]   # route_idx -> [stop_indices in order]
    stop_route = list(range(n))           # stop_idx  -> route_idx

    def route_vol(r_idx):
        return sum(stops[k]['vol'] for k in routes[r_idx]) if routes[r_idx] else 0.0

    for _saving, i, j in savings:
        ri = stop_route[i]
        rj = stop_route[j]

        # Skip if already merged or route dissolved
        if ri == rj or routes[ri] is None or routes[rj] is None:
            continue

        ri_r = routes[ri]
        rj_r = routes[rj]

        # Both stops must be endpoints of their respective routes
        ri_first, ri_last = ri_r[0], ri_r[-1]
        rj_first, rj_last = rj_r[0], rj_r[-1]

        if   i == ri_last  and j == rj_first: new_route = ri_r + rj_r
        elif i == ri_first and j == rj_last:  new_route = rj_r + ri_r
        elif i == ri_last  and j == rj_last:  new_route = ri_r + list(reversed(rj_r))
        elif i == ri_first and j == rj_first: new_route = list(reversed(ri_r)) + rj_r
        else:
            continue    # interior stops — cannot merge at this edge

        # Capacity and drop-count feasibility vs. the target truck size
        combined_vol = route_vol(ri) + route_vol(rj)
        if combined_vol > target_cap * cap_tolerance:
            continue    # exceeds target truck capacity — don't merge
        if len(new_route) > 10:
            continue    # 5MT / 2.5MT max-drops guard

        # Perform the merge: keep route ri, dissolve rj
        routes[ri] = new_route
        routes[rj] = None
        for k in new_route:
            stop_route[k] = ri

    # Return stop-dict lists, skipping dissolved routes
    return [[stops[k] for k in route] for route in routes if route is not None]


# ─────────────────────────────────────────────────────────────
#  MAIN ROUTING ENGINE  (Clarke-Wright edition)
# ─────────────────────────────────────────────────────────────
def run_routing_engine(orders_df) -> tuple:
    col_map = normalize_headers(orders_df.columns.tolist())

    # ── Step 1: Build consolidated stops ──────────────────────
    stops_dict: dict = {}
    for _, row in orders_df.iterrows():
        row_dict = row.to_dict()
        vol = 0
        try:
            vol = float(safe_get(row_dict, col_map, 'tfor_qty', 0) or
                        safe_get(row_dict, col_map, 'do_qty', 0) or 0)
        except Exception:
            pass
        if vol <= 0:
            continue

        cust = str(safe_get(row_dict, col_map, 'customer_name', '')).strip() or 'UNKNOWN'
        addr = str(safe_get(row_dict, col_map, 'shipping_address', '')).strip() or cust

        doc_num = str(safe_get(row_dict, col_map, 'so_number', '')).strip().upper()
        is_stock_transfer = doc_num.startswith('TFORPFC')

        addr_key = re.sub(r'\s+', ' ', addr.upper().strip())
        loc_raw  = str(safe_get(row_dict, col_map, 'location', 'NA'))
        wh       = parse_location_cell(loc_raw)
        key      = addr_key + '|||' + wh

        cluster = str(safe_get(row_dict, col_map, 'cluster_id', 'UN')).strip().upper()
        area    = str(safe_get(row_dict, col_map, 'area', ''))

        try:
            lat = float(safe_get(row_dict, col_map, 'delivery_latitude',  0) or 0)
            lng = float(safe_get(row_dict, col_map, 'delivery_longitude', 0) or 0)
        except Exception:
            lat, lng = 0.0, 0.0

        s = serialise_row(row_dict)
        s['_pfc_doc_num'] = doc_num
        s['_pfc_vol'] = vol

        if key not in stops_dict:
            stops_dict[key] = {
                'rows': [s], 'vol': vol, 'cluster_id': cluster,
                'wh': wh, 'wh_group': get_wh_group(wh),
                'match_area': clean_area_name(area),
                'lat': lat, 'lng': lng,
                'customer_name': cust, 'shipping_address': addr, 'area': area,
                'all_customers': [cust],
                'is_stock_transfer': is_stock_transfer,
                'doc_number': doc_num,
            }
        else:
            stops_dict[key]['vol'] += vol
            stops_dict[key]['rows'].append(s)
            if cust not in stops_dict[key]['all_customers']:
                stops_dict[key]['all_customers'].append(cust)
            stops_dict[key]['customer_name'] = ' / '.join(stops_dict[key]['all_customers'])
            if is_stock_transfer:
                stops_dict[key]['is_stock_transfer'] = True

    all_stops = list(stops_dict.values())

    # ── Step 1.5: Resolve lat/lng from shipping address ────────
    import urllib.request as _ureq, urllib.parse as _uparse, time as _gtime
    conn_geo = open_db()
    c_geo    = conn_geo.cursor()

    for stop in all_stops:
        addr = stop.get('shipping_address', '').strip()
        cache_key = re.sub(r'\s+', ' ', addr.upper().strip()) if addr else ''

        # Priority 1: coords from the uploaded file (delivery_latitude/longitude columns)
        if stop.get('lat') and stop.get('lng'):
            # Seed the cache so future lookups by address are instant
            if cache_key:
                c_geo.execute(
                    'INSERT OR IGNORE INTO address_fwd_cache (address_key,lat,lng,display_name) VALUES (?,?,?,?)',
                    (cache_key, stop['lat'], stop['lng'], addr))
                conn_geo.commit()
            continue  # no geocoding needed

        if not addr:
            continue

        # Priority 2: address cache (previously geocoded or seeded from file)
        c_geo.execute('SELECT lat, lng FROM address_fwd_cache WHERE address_key=?', (cache_key,))
        row = c_geo.fetchone()
        if row and row[0] and row[1]:
            stop['lat'] = float(row[0])
            stop['lng'] = float(row[1])
            continue

        # Priority 3: Nominatim (only when no coords in file and not cached)
        for attempt in _address_candidates(addr):
            try:
                hit = _nominatim_search(attempt)
                _gtime.sleep(1.05)
                if hit:
                    new_lat = round(float(hit['lat']), 6)
                    new_lng = round(float(hit['lon']), 6)
                    disp    = hit.get('display_name', '')
                    c_geo.execute(
                        '''INSERT OR REPLACE INTO address_fwd_cache
                           (address_key, lat, lng, display_name) VALUES (?,?,?,?)''',
                        (cache_key, new_lat, new_lng, disp))
                    conn_geo.commit()
                    stop['lat'] = new_lat
                    stop['lng'] = new_lng
                    break
            except Exception:
                _gtime.sleep(1.05)
                continue

    conn_geo.close()

    # ── Step 2: Separate by wh_group (hard WH boundary) ───────
    # Float stops (RGM etc.) go to unrouted basket — planner assigns manually
    floated_stops = [s for s in all_stops if s['wh_group'] in FLOAT_WH_GROUPS]
    by_wh_group: dict = defaultdict(list)
    for stop in all_stops:
        if stop['wh_group'] not in FLOAT_WH_GROUPS:
            by_wh_group[stop['wh_group']].append(stop)

    trucks_out = []
    truck_idx  = 0

    for wh_group, group_stops in by_wh_group.items():
        wh_code = next(
            (code for code, d in WAREHOUSE_DATA.items() if d['group'] == wh_group),
            'FCSC'
        )
        wh_lat, wh_lng = get_wh_coords(wh_code)

        # ── Clarke-Wright Savings → build routes ──────────────
        # Stops heavier than a 5MT truck run solo (they'll get an
        # appropriately-sized truck type assigned below).
        STD_CAP = 5000
        std_stops   = [s for s in group_stops if s['vol'] <= STD_CAP * SETTINGS['cap_tolerance']]
        heavy_stops = [s for s in group_stops if s['vol']  > STD_CAP * SETTINGS['cap_tolerance']]

        routes = clarke_wright_routing(wh_lat, wh_lng, std_stops,
                                       target_cap=STD_CAP)
        # Add each oversized stop as its own single-stop route
        for hs in heavy_stops:
            routes.append([hs])

        for truck_stops in routes:
            truck_vol = sum(s['vol'] for s in truck_stops)

            # ── Truck type ────────────────────────────────────
            v_type = best_truck_type(truck_vol, len(truck_stops))
            if v_type['label'] == '2.5MT' and truck_vol > SETTINGS['prefer_5mt_min_vol']:
                v_type = TRUCK_TYPES[1]

            # ── TSP reorder from warehouse ─────────────────────
            truck_stops = nearest_neighbor_tsp(wh_lat, wh_lng, truck_stops)

            # ── Total route distance ───────────────────────────
            total_km = calculate_total_km(wh_lat, wh_lng, truck_stops)

            # ── Primary WH and WH list ─────────────────────────
            wh_vols: dict = {}
            for s in truck_stops:
                wh_vols[s['wh']] = wh_vols.get(s['wh'], 0) + s['vol']
            primary_wh = max(wh_vols, key=wh_vols.get) if wh_vols else wh_code
            all_whs    = sorted(wh_vols.keys())

            truck_idx += 1
            truck_id   = f'TRK-{truck_idx:03d}'
            rated_cap  = v_type['cap']
            util_pct   = round(truck_vol / rated_cap, 4)

            truck = {
                'truck_id':    truck_id,
                'truck_type':  v_type['label'],
                'truck_cap':   v_type['cap'],
                'rated_cap':   rated_cap,
                'acc_weight':  round(truck_vol, 3),
                'util_pct':    util_pct,
                'pickup_wh':   primary_wh,
                'all_whs':     all_whs,
                'wh_group':    wh_group,
                'trucker_code': '',
                'ref_number':   '',
                'total_km':    total_km,
                'stops':       [],
            }

            for seq_idx, stop in enumerate(truck_stops):
                seq = seq_idx + 1
                if v_type['label'] == '5MT':
                    add_fee = SETTINGS['add_drop_fee'] if seq > SETTINGS['add_drop_threshold_5mt'] else 0
                elif v_type['label'] == '2.5MT':
                    add_fee = SETTINGS['add_drop_fee'] if seq > SETTINGS['add_drop_threshold_25mt'] else 0
                else:
                    add_fee = 0

                truck['stops'].append({
                    'seq':              seq,
                    'cluster_id':       stop['cluster_id'],
                    'customer_name':    stop['customer_name'],
                    'shipping_address': stop['shipping_address'],
                    'area':             stop['area'],
                    'match_area':       stop['match_area'],
                    'lat':              stop['lat'],
                    'lng':              stop['lng'],
                    'vol':              round(stop['vol'], 3),
                    'wh':               stop['wh'],
                    'add_drop_fee':         add_fee,
                    'is_stock_transfer':    stop.get('is_stock_transfer', False),
                    'doc_number':           stop.get('doc_number', ''),
                    'all_customers':        stop.get('all_customers', [stop['customer_name']]),
                    'rows':                 stop['rows'],
                })

            trucks_out.append(truck)

    return trucks_out, floated_stops

# ─────────────────────────────────────────────────────────────
#  DATABASE
# ─────────────────────────────────────────────────────────────
def init_db():
    conn = open_db()
    c = conn.cursor()

    # WAL + busy_timeout already set by open_db()
    c.execute('''CREATE TABLE IF NOT EXISTS route_plans (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        plan_date  TEXT NOT NULL,
        plan_name  TEXT NOT NULL,
        created_at TEXT DEFAULT (datetime('now','localtime')),
        created_by TEXT DEFAULT '',
        status     TEXT DEFAULT 'draft',
        data       TEXT NOT NULL,
        summary    TEXT
    )''')
    # Migration: add created_by if missing
    try:
        c.execute("ALTER TABLE route_plans ADD COLUMN created_by TEXT DEFAULT ''")
    except Exception:
        pass

    # ── Shared live working plan (one row per date, all planners share it) ──
    c.execute('''CREATE TABLE IF NOT EXISTS working_plans (
        plan_date  TEXT PRIMARY KEY,
        data       TEXT NOT NULL,
        updated_by TEXT DEFAULT '',
        updated_at TEXT DEFAULT (datetime('now','localtime'))
    )''')
    # ────────────────────────────────────────────────────────────────────────
    # Migration: add shipping_address to monitoring_records if missing
    try:
        c.execute("ALTER TABLE monitoring_records ADD COLUMN shipping_address TEXT DEFAULT ''")
    except Exception:
        pass

    c.execute('''CREATE TABLE IF NOT EXISTS monitoring_records (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        plan_id       INTEGER,
        plan_date     TEXT,
        truck_id      TEXT,
        truck_type    TEXT,
        seq           INTEGER,
        customer_name TEXT,
        cluster_id    TEXT,
        area          TEXT,
        status        TEXT,
        receiving_time TEXT,
        actual_done   TEXT,
        otif_status   TEXT,
        concerns      TEXT,
        remarks       TEXT,
        trucker_code  TEXT,
        shipping_address   TEXT DEFAULT '',
        food_safety_issue  TEXT DEFAULT 'NONE',
        food_safety_detail TEXT DEFAULT '',
        crew_issue         TEXT DEFAULT 'NONE',
        crew_detail        TEXT DEFAULT '',
        saved_at      TEXT DEFAULT (datetime('now','localtime'))
    )''')
    # Migrations: add any columns that may be missing in older DBs
    _mon_cols = [
        ("plan_id",           "INTEGER"),
        ("plan_date",         "TEXT"),
        ("truck_type",        "TEXT"),
        ("cluster_id",        "TEXT"),
        ("area",              "TEXT"),
        ("receiving_time",    "TEXT"),
        ("actual_done",       "TEXT"),
        ("otif_status",       "TEXT"),
        ("concerns",          "TEXT"),
        ("remarks",           "TEXT"),
        ("trucker_code",      "TEXT"),
        ("shipping_address",  "TEXT DEFAULT ''"),
        ("food_safety_issue", "TEXT DEFAULT 'NONE'"),
        ("food_safety_detail","TEXT DEFAULT ''"),
        ("crew_issue",        "TEXT DEFAULT 'NONE'"),
        ("crew_detail",       "TEXT DEFAULT ''"),
        ("do_number",         "TEXT DEFAULT ''"),
        ("return_date",       "TEXT DEFAULT ''"),
        ("pod_remarks",       "TEXT DEFAULT ''"),
        ("saved_at",          "TEXT DEFAULT (datetime('now','localtime'))"),
    ]
    for col, coltype in _mon_cols:
        try:
            c.execute(f"ALTER TABLE monitoring_records ADD COLUMN {col} {coltype}")
        except Exception:
            pass  # column already exists

    # Cold Chain compliance records
    c.execute('''CREATE TABLE IF NOT EXISTS cold_chain_records (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        plan_date     TEXT NOT NULL,
        truck_id      TEXT NOT NULL,
        do_number     TEXT DEFAULT '',
        customer_name TEXT DEFAULT '',
        has_issue     INTEGER DEFAULT 0,
        issue_details TEXT DEFAULT '',
        saved_at      TEXT DEFAULT (datetime('now','localtime'))
    )''')

    # POD return compliance records
    c.execute('''CREATE TABLE IF NOT EXISTS pod_records (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        do_number     TEXT NOT NULL,
        customer_name TEXT DEFAULT '',
        truck_id      TEXT DEFAULT '',
        delivery_date TEXT NOT NULL,
        return_date   TEXT DEFAULT '',
        days_aging    INTEGER DEFAULT NULL,
        status        TEXT DEFAULT 'PENDING',
        remarks       TEXT DEFAULT '',
        saved_at      TEXT DEFAULT (datetime('now','localtime'))
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS geocode_cache (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        lat          REAL NOT NULL,
        lng          REAL NOT NULL,
        city         TEXT,
        municipality TEXT,
        province     TEXT,
        display_name TEXT,
        cached_at    TEXT DEFAULT (datetime('now','localtime')),
        UNIQUE(lat, lng)
    )''')

    # Forward geocoding cache: address string → lat/lng
    c.execute('''CREATE TABLE IF NOT EXISTS address_fwd_cache (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        address_key  TEXT NOT NULL UNIQUE,
        lat          REAL,
        lng          REAL,
        display_name TEXT,
        confidence   REAL DEFAULT 1.0,
        cached_at    TEXT DEFAULT (datetime('now','localtime'))
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS rate_master (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        trucker     TEXT NOT NULL,
        capacity_kg INTEGER NOT NULL,
        pickup_wh   TEXT NOT NULL,
        area        TEXT NOT NULL,
        rate        REAL DEFAULT 0,
        cost_per_kg REAL DEFAULT 0,
        is_active   INTEGER DEFAULT 1,
        updated_at  TEXT DEFAULT (datetime('now','localtime'))
    )''')

    conn.commit()
    conn.close()

# ─────────────────────────────────────────────────────────────
#  FLASK ROUTES — Core
# ─────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/warehouses', methods=['GET'])
def get_warehouses():
    return jsonify(WAREHOUSE_DATA)

@app.route('/api/truck-types', methods=['GET'])
def get_truck_types():
    return jsonify(TRUCK_TYPES)

@app.route('/api/upload', methods=['POST'])
def upload_file():
    router = request.args.get('router', 'default')
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'error': 'No file selected'}), 400

    filename = secure_filename(f.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    f.save(filepath)

    pd = _get_pd()
    df = None
    try:
        if filename.lower().endswith(('.xlsx', '.xls', '.xlsm')):
            xl = pd.ExcelFile(filepath)
            target = 'Order Details' if 'Order Details' in xl.sheet_names else xl.sheet_names[0]
            df = pd.read_excel(filepath, sheet_name=target)
        else:
            try:
                df = pd.read_csv(filepath, encoding='utf-8')
            except Exception:
                df = pd.read_csv(filepath, encoding='latin1')
    except Exception as e:
        return jsonify({'error': f'Cannot parse file: {str(e)}'}), 400

    df = df.dropna(how='all')
    df.columns = [str(c).strip() for c in df.columns]
    col_map = normalize_headers(df.columns.tolist())

    # Accept either SO TFOR Qty (col N) or DO Qty (col O)
    vol_key = 'tfor_qty' if 'tfor_qty' in col_map else 'do_qty'
    missing = [r for r in ('customer_name',) if r not in col_map]
    if missing or vol_key not in col_map:
        miss2 = missing + ([] if vol_key in col_map else ['SO TFOR Qty / DO Qty'])
        return jsonify({'error': f'Could not find required columns: {miss2}. '
                                 f'Found: {df.columns.tolist()}'}), 400

    vol_col  = col_map.get(vol_key)
    cust_col = col_map.get('customer_name')
    ship_col = col_map.get('ship_date')
    total_vol   = float(df[vol_col].sum()) if vol_col else 0
    unique_cust = int(df[cust_col].nunique()) if cust_col else 0
    ship_dates  = []
    if ship_col:
        ship_dates = df[ship_col].dropna().astype(str).unique()[:3].tolist()
    records = [serialise_row(r.to_dict()) for _, r in df.iterrows()]
    with open(get_router_temp('orders', router), 'w') as fp:
        json.dump({'records': records, 'col_map': col_map}, fp)
    return jsonify({
        'success': True, 'total_orders': len(df), 'total_volume': round(total_vol,2),
        'unique_customers': unique_cust, 'columns': df.columns.tolist(),
        'col_map': col_map, 'ship_dates': ship_dates, 'filename': filename,
    })


@app.route('/api/run-routing', methods=['POST'])
def run_routing():
    router    = request.args.get('router', 'default')
    plan_date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    temp_orders = get_router_temp('orders', router)
    temp_plan   = get_router_temp('plan', router)
    if not os.path.exists(temp_orders):
        return jsonify({'error': 'No orders loaded. Please upload a file first.'}), 400
    try:
        with open(temp_orders) as fp:
            payload = json.load(fp)
        pd = _get_pd()
        df = pd.DataFrame(payload['records'])
        trucks, floated = run_routing_engine(df)
        with open(temp_plan, 'w') as fp:
            json.dump(trucks, fp, default=str)
        # ── Write to shared live plan so all browsers see the new plan ──
        set_working_plan(plan_date, trucks, router)
        total_drops = sum(len(t['stops']) for t in trucks)
        avg_util    = (sum(t['util_pct'] for t in trucks) / len(trucks)) if trucks else 0
        total_vol   = sum(t['acc_weight'] for t in trucks)
        under_60 = sum(1 for t in trucks if t['util_pct'] < 0.60)
        range_60 = sum(1 for t in trucks if 0.60 <= t['util_pct'] < 0.80)
        above_80 = sum(1 for t in trucks if t['util_pct'] >= 0.80)
        return jsonify({'success': True, 'truck_count': len(trucks), 'total_drops': total_drops,
                        'total_volume': round(total_vol,2), 'avg_utilization': round(avg_util*100,1),
                        'util_breakdown': {'under_60':under_60,'range_60_80':range_60,'above_80':above_80},
                        'trucks': trucks, 'floated_stops': floated})
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/api/get-plan', methods=['GET'])
def get_plan():
    router    = request.args.get('router', 'default')
    plan_date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    temp_plan = get_router_temp('plan', router)
    # Prefer shared working plan (authoritative) over local temp file
    trucks, _, _ = get_working_plan(plan_date)
    if trucks is not None:
        # Keep temp file in sync so other endpoints that read it work correctly
        with open(temp_plan, 'w') as fp:
            json.dump(trucks, fp, default=str)
        return jsonify({'success': True, 'trucks': trucks})
    # Fall back to local temp file if no shared plan yet
    if not os.path.exists(temp_plan):
        return jsonify({'success': False, 'trucks': []})
    with open(temp_plan) as fp:
        trucks = json.load(fp)
    return jsonify({'success': True, 'trucks': trucks})


@app.route('/api/update-plan', methods=['POST'])
def update_plan():
    router    = request.args.get('router', 'default')
    plan_date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    temp_plan = get_router_temp('plan', router)
    data      = request.get_json()
    if not data or 'trucks' not in data:
        return jsonify({'error': 'Invalid payload'}), 400
    for t in data['trucks']:
        vol       = sum(s['vol'] for s in t['stops'])
        rated_cap = float(t.get('rated_cap') or t.get('truck_cap') or 1)
        t['acc_weight'] = round(vol, 3)
        t['util_pct']   = round(vol / rated_cap, 4) if rated_cap else t.get('util_pct', 0)
        # Recalculate total_km based on current stop order
        wh_lat, wh_lng = get_wh_coords(t.get('pickup_wh', 'FCSC'))
        t['total_km'] = calculate_total_km(wh_lat, wh_lng, t['stops'])
        # Refresh seq numbers and add_drop_fee
        vtype_label = t.get('truck_type', '')
        for i, s in enumerate(t['stops']):
            s['seq'] = i + 1
            seq = i + 1
            if vtype_label == '5MT':
                s['add_drop_fee'] = SETTINGS['add_drop_fee'] if seq > SETTINGS['add_drop_threshold_5mt'] else 0
            elif vtype_label == '2.5MT':
                s['add_drop_fee'] = SETTINGS['add_drop_fee'] if seq > SETTINGS['add_drop_threshold_25mt'] else 0
            else:
                s['add_drop_fee'] = 0
    with open(temp_plan, 'w') as fp:
        json.dump(data['trucks'], fp, default=str)
    version = os.path.getmtime(temp_plan)
    # ── Broadcast to all other browsers via shared live plan ──
    set_working_plan(plan_date, data['trucks'], router)
    return jsonify({'success': True, 'trucks': data['trucks'], 'version': version})


@app.route('/api/plan-state', methods=['GET'])
def get_plan_state():
    """Lightweight collab poll — checks shared working_plans table for the given date.
    Returns updated_at timestamp and who last changed it so browsers can auto-apply."""
    plan_date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    try:
        conn = open_db(); c = conn.cursor()
        c.execute('SELECT updated_at, updated_by, data FROM working_plans WHERE plan_date=?', (plan_date,))
        row = c.fetchone(); conn.close()
        if not row:
            return jsonify({'version': None, 'exists': False, 'updated_by': '', 'updated_at': ''})
        trucks = []
        try: trucks = json.loads(row[2])
        except Exception: pass
        return jsonify({
            'version':     row[0],
            'updated_at':  row[0],
            'updated_by':  row[1] or '',
            'exists':      True,
            'truck_count': len(trucks),
            'total_stops': sum(len(t.get('stops', [])) for t in trucks),
        })
    except Exception as e:
        return jsonify({'version': None, 'exists': False, 'updated_by': '', 'error': str(e)})


@app.route('/api/working-plan', methods=['GET'])
def get_working_plan_endpoint():
    """Return the full shared live plan for a date.  Called when a browser needs to sync."""
    plan_date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    trucks, updated_by, updated_at = get_working_plan(plan_date)
    if trucks is None:
        return jsonify({'trucks': [], 'updated_by': '', 'updated_at': '', 'exists': False})
    return jsonify({'trucks': trucks, 'updated_by': updated_by,
                    'updated_at': updated_at, 'exists': True})


# ─────────────────────────────────────────────────────────────
#  PER-TRUCK PDF ROUTE SHEET
# ─────────────────────────────────────────────────────────────
@app.route('/api/download-truck-pdf', methods=['GET', 'POST'])
def download_truck_pdf():
    try:
        import io, tempfile, base64
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        import matplotlib.patches as mpatches
        from fpdf import FPDF

        # Support both GET (fallback/legacy) and POST (with browser screenshot)
        if request.method == 'POST':
            body     = request.get_json() or {}
            router   = body.get('router', 'default')
            truck_id = body.get('truck_id', '')
            map_b64  = body.get('map_image_b64', '')  # data:image/png;base64,...
        else:
            router   = request.args.get('router', 'default')
            truck_id = request.args.get('truck_id', '')
            map_b64  = ''
        temp_plan = get_router_temp('plan', router)
        if not os.path.exists(temp_plan):
            return jsonify({'error': 'No plan loaded - generate a route first'}), 400
        with open(temp_plan) as fp:
            trucks = json.load(fp)
        truck = next((t for t in trucks if t['truck_id'] == truck_id), None)
        if not truck:
            return jsonify({'error': f'{truck_id} not found in current plan'}), 404

        # ── Warehouse info ────────────────────────────────────
        wh_code = truck.get('pickup_wh', '')
        wh_data = WAREHOUSE_DATA.get(wh_code, {})
        wh_name = wh_data.get('name', wh_code or 'Warehouse')
        wh_lat  = wh_data.get('lat', 0)
        wh_lng  = wh_data.get('lng', 0)
        wh_maps = f'https://maps.google.com/?q={wh_lat},{wh_lng}'
        today_str = datetime.now().strftime('%B %d, %Y')

        # ─────────────────────────────────────────────────────
        #  1. ROUTE MAP IMAGE (matplotlib schematic)
        # ─────────────────────────────────────────────────────
        truck_stops = truck.get('stops', [])

        def _draw_route_map():
            # ── Primary: real OSM tile map via staticmap ──────────────────────
            try:
                from staticmap import StaticMap, CircleMarker, Line as SMLine

                # Build ordered coord list: warehouse first, then stops in sequence
                coords = [(wh_lng, wh_lat)]
                for s in truck_stops:
                    slng = s.get('lng') or wh_lng
                    slat = s.get('lat') or wh_lat
                    coords.append((slng, slat))

                # Build StaticMap — headers param only exists in newer versions
                OSM = 'https://tile.openstreetmap.org/{z}/{x}/{y}.png'
                try:
                    sm = StaticMap(680, 360, url_template=OSM,
                                   headers={'User-Agent': 'PFC-Logistics/1.0'})
                except TypeError:
                    sm = StaticMap(680, 360, url_template=OSM)

                # Route line (blue)
                sm.add_line(SMLine(coords, '#1a56db', 3))

                # Warehouse marker (green, larger)
                sm.add_marker(CircleMarker(coords[0], '#16a34a', 16))

                # Stop markers (red)
                for c in coords[1:]:
                    sm.add_marker(CircleMarker(c, '#e53935', 11))

                img = sm.render()

                # Overlay numbered labels on the tile image using Pillow
                try:
                    from PIL import ImageDraw, ImageFont
                    draw = ImageDraw.Draw(img)
                    img_w, img_h = img.size

                    # Helper: project lat/lng to pixel (uses the rendered image bounds)
                    # staticmap renders with a zoom auto-selected; we re-derive pixels
                    # by checking where our coords landed vs the image dimensions.
                    # Simpler: draw labels at approximate relative positions.
                    all_lngs = [c[0] for c in coords]
                    all_lats = [c[1] for c in coords]
                    lng_min, lng_max = min(all_lngs), max(all_lngs)
                    lat_min, lat_max = min(all_lats), max(all_lats)
                    lng_span = max(lng_max - lng_min, 0.001)
                    lat_span = max(lat_max - lat_min, 0.001)
                    pad_px = 40  # matches staticmap's internal padding roughly

                    def latlon_to_px(lng, lat):
                        fx = (lng - lng_min) / lng_span
                        fy = 1.0 - (lat - lat_min) / lat_span  # y flipped
                        px = int(pad_px + fx * (img_w - 2 * pad_px))
                        py = int(pad_px + fy * (img_h - 2 * pad_px))
                        return px, py

                    try:
                        font = ImageFont.truetype("arial.ttf", 11)
                        font_sm = ImageFont.truetype("arial.ttf", 9)
                    except Exception:
                        font = ImageFont.load_default()
                        font_sm = font

                    # Warehouse label
                    wx, wy = latlon_to_px(wh_lng, wh_lat)
                    short_wh = wh_name.split('(')[0].strip()[:18]
                    draw.rectangle([wx+9, wy-14, wx+9+len(short_wh)*6+4, wy+2],
                                   fill='#d1fae5', outline='#16a34a')
                    draw.text((wx+11, wy-13), short_wh, fill='#064e3b', font=font_sm)

                    # Stop number labels
                    for i, stop in enumerate(truck_stops):
                        slng = stop.get('lng') or wh_lng
                        slat = stop.get('lat') or wh_lat
                        sx, sy = latlon_to_px(slng, slat)
                        label = str(stop.get('seq', i + 1))
                        draw.ellipse([sx-8, sy-8, sx+8, sy+8], fill='#e53935', outline='white')
                        draw.text((sx-4 if len(label)==1 else sx-6, sy-7),
                                  label, fill='white', font=font_sm)

                except Exception:
                    pass  # Labels optional — plain tile map still fine

                buf = io.BytesIO()
                img.save(buf, format='PNG')
                buf.seek(0)
                return buf.read()

            except Exception:
                pass  # Fall through to matplotlib schematic

            # ── Fallback: matplotlib schematic (no internet needed) ───────────
            all_lats = [wh_lat] + [s.get('lat') or wh_lat for s in truck_stops]
            all_lngs = [wh_lng] + [s.get('lng') or wh_lng for s in truck_stops]
            lat_min, lat_max = min(all_lats), max(all_lats)
            lng_min, lng_max = min(all_lngs), max(all_lngs)
            pad_lat = max((lat_max - lat_min) * 0.20, 0.025)
            pad_lng = max((lng_max - lng_min) * 0.20, 0.025)

            fig, ax = plt.subplots(figsize=(7.8, 3.6))
            fig.patch.set_facecolor('#e8f0fe')
            ax.set_facecolor('#dce8f8')
            ax.grid(True, color='#b8d0f0', linewidth=0.4, linestyle='--', alpha=0.6)
            ax.set_xlim(lng_min - pad_lng, lng_max + pad_lng)
            ax.set_ylim(lat_min - pad_lat, lat_max + pad_lat)
            ax.tick_params(labelsize=6, colors='#666')
            ax.set_xlabel('Longitude', fontsize=6.5, color='#555')
            ax.set_ylabel('Latitude',  fontsize=6.5, color='#555')

            route_lngs = [wh_lng] + [s.get('lng') or wh_lng for s in truck_stops]
            route_lats = [wh_lat] + [s.get('lat') or wh_lat for s in truck_stops]
            ax.plot(route_lngs, route_lats, '-', color='#1a56db',
                    linewidth=2.0, alpha=0.80, zorder=2)

            for i in range(len(route_lngs) - 1):
                dx = route_lngs[i+1] - route_lngs[i]
                dy = route_lats[i+1] - route_lats[i]
                mx, my = route_lngs[i] + dx*0.5, route_lats[i] + dy*0.5
                ax.annotate('',
                    xy=(mx + dx*0.01, my + dy*0.01),
                    xytext=(mx - dx*0.01, my - dy*0.01),
                    arrowprops=dict(arrowstyle='->', color='#1a56db', lw=1.4),
                    zorder=3)

            ax.scatter([wh_lng], [wh_lat], s=200, c='#16a34a', zorder=6,
                       marker='*', edgecolors='white', linewidths=0.9)
            short_wh = wh_name.split('(')[0].strip()[:20]
            ax.annotate(f'START\n{short_wh}',
                xy=(wh_lng, wh_lat), xytext=(6, 6), textcoords='offset points',
                fontsize=5.5, color='#064e3b', fontweight='bold',
                bbox=dict(boxstyle='round,pad=0.25', fc='#d1fae5', ec='#16a34a', lw=0.8))

            for i, stop in enumerate(truck_stops):
                slat = stop.get('lat') or wh_lat
                slng = stop.get('lng') or wh_lng
                ax.scatter([slng], [slat], s=90, c='#e53935', zorder=5,
                           edgecolors='white', linewidths=0.9)
                cust = (stop.get('customer_name') or '')[:18]
                ax.annotate(f"{stop.get('seq', i+1)}. {cust}",
                    xy=(slng, slat), xytext=(5, 4), textcoords='offset points',
                    fontsize=5, color='#111',
                    bbox=dict(boxstyle='round,pad=0.22', fc='white',
                              ec='#e53935', lw=0.6, alpha=0.88))

            wh_p   = mpatches.Patch(color='#16a34a', label='Warehouse (Start)')
            stop_p = mpatches.Patch(color='#e53935', label=f'Drops ({len(truck_stops)})')
            ax.legend(handles=[wh_p, stop_p], loc='lower right',
                      fontsize=6, framealpha=0.92)
            ax.set_title(
                f'{truck_id}  |  {truck.get("truck_type","")}  |  '
                f'{truck.get("acc_weight",0):,.0f} kg  |  {len(truck_stops)} drops  |  Est. {truck.get("total_km",0)} km',
                fontsize=7.5, color='#002d62', fontweight='bold', pad=5)

            plt.tight_layout(pad=0.4)
            buf = io.BytesIO()
            fig.savefig(buf, format='PNG', dpi=140, bbox_inches='tight')
            plt.close(fig)
            buf.seek(0)
            return buf.read()

        # Use browser screenshot if provided, otherwise render server-side
        if map_b64 and map_b64.startswith('data:image'):
            try:
                header, encoded = map_b64.split(',', 1)
                map_png = base64.b64decode(encoded)
            except Exception:
                map_png = _draw_route_map()
        else:
            map_png = _draw_route_map()

        # ─────────────────────────────────────────────────────
        #  2. BUILD PDF
        # ─────────────────────────────────────────────────────
        class PDF(FPDF):
            def header(self):
                self.set_font('Helvetica', 'B', 8)
                self.set_text_color(120, 120, 120)
                self.cell(0, 5, f'PFC Logistics - Confidential Route Sheet - {today_str}', align='R')
                self.ln(3)
            def footer(self):
                self.set_y(-12)
                self.set_font('Helvetica', 'I', 7.5)
                self.set_text_color(150, 150, 150)
                self.cell(0, 5, f'Page {self.page_no()} - {truck_id} Route Sheet', align='C')

        pdf = PDF(orientation='P', unit='mm', format='A4')
        pdf.set_auto_page_break(auto=True, margin=16)
        pdf.set_margins(12, 14, 12)
        pdf.add_page()

        # Title bar
        pdf.set_fill_color(0, 45, 98)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('Helvetica', 'B', 14)
        pdf.cell(0, 11, 'PFC LOGISTICS - DELIVERY ROUTE SHEET', fill=True, ln=True, align='C')

        # Truck summary bar
        util_pct = round((truck.get('util_pct') or 0) * 100, 1)
        stops_n  = len(truck_stops)
        pdf.set_fill_color(0, 78, 152)
        pdf.set_font('Helvetica', 'B', 9.5)
        pdf.cell(0, 7,
            f"  {truck_id}   |   {truck.get('truck_type','')}   |   "
            f"{truck.get('acc_weight',0):,.1f} kg / {truck.get('truck_cap',0):,} kg   |   "
            f"{util_pct}% util   |   {stops_n} drops",
            fill=True, ln=True)

        pdf.set_text_color(0, 0, 0)
        pdf.set_font('Helvetica', '', 8)
        pdf.set_fill_color(232, 240, 254)
        extras = []
        if truck.get('trucker_code'): extras.append(f"Trucker: {truck['trucker_code']}")
        if truck.get('ref_number'):   extras.append(f"Ref #: {truck['ref_number']}")
        if truck.get('total_km'):     extras.append(f"Est. Distance: {truck['total_km']} km")
        extras.append(f"Pickup WH: {wh_name}")
        pdf.cell(0, 6, '  ' + '   |   '.join(extras), fill=True, ln=True, border=1)

        # ── Route map image ────────────────────────────────
        pdf.ln(3)
        map_tmp = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
        map_tmp.write(map_png); map_tmp.flush(); map_tmp.close()
        usable_w = pdf.w - pdf.l_margin - pdf.r_margin
        map_h = usable_w * (3.6 / 7.8)   # keep aspect ratio
        pdf.image(map_tmp.name, x=pdf.l_margin, y=pdf.get_y(), w=usable_w, h=map_h)
        os.unlink(map_tmp.name)
        pdf.set_y(pdf.get_y() + map_h + 3)

        # ── Drop table header ──────────────────────────────
        CW = [10, 55, 68, 24, 29]   # #, Customer, Address, Vol kg, Map
        pdf.set_fill_color(0, 45, 98)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('Helvetica', 'B', 8)
        for w, h in zip(CW, ['#', 'Customer / Consignee', 'Delivery Address', 'Vol (kg)', 'Map Link']):
            pdf.multi_cell(w, 5, h, border=1, fill=True, align='C',
                           max_line_height=4, new_x='RIGHT', new_y='TOP')
        pdf.ln(10)

        # ── Warehouse row ──────────────────────────────────
        pdf.set_fill_color(220, 234, 255)
        pdf.set_text_color(0, 45, 98)
        pdf.set_font('Helvetica', 'B', 8)
        pdf.cell(CW[0], 8, 'WH',   border=1, fill=True, align='C')
        pdf.cell(CW[1], 8, 'WAREHOUSE PICKUP (START)', border=1, fill=True)
        wn = wh_name if len(wh_name) <= 42 else wh_name[:40] + '..'
        pdf.cell(CW[2], 8, wn,     border=1, fill=True)
        pdf.cell(CW[3], 8, '---',  border=1, fill=True, align='C')
        pdf.set_font('Helvetica', 'BU', 8)
        pdf.set_text_color(0, 80, 200)
        pdf.cell(CW[4], 8, 'Open Map', border=1, align='C', link=wh_maps, fill=True)
        pdf.set_text_color(0, 0, 0)
        pdf.ln()

        # ── Stop rows ──────────────────────────────────────
        for idx, stop in enumerate(truck_stops):
            even = (idx % 2 == 0)
            pdf.set_fill_color(248, 250, 255) if even else pdf.set_fill_color(255, 255, 255)
            lat = stop.get('lat') or 0
            lng = stop.get('lng') or 0
            if lat and lng:
                maps_url = f'https://maps.google.com/?q={lat},{lng}'
            else:
                q = (stop.get('shipping_address') or '').replace(' ', '+')
                maps_url = f'https://maps.google.com/?q={q}'

            cust = stop.get('customer_name', '') or ''
            addr = stop.get('shipping_address', '') or ''
            cust_s = (cust[:32] + '..') if len(cust) > 32 else cust
            addr_s = (addr[:42] + '..') if len(addr) > 42 else addr
            vol_s  = f"{stop.get('vol', 0):,.1f}"

            pdf.set_font('Helvetica', '', 8)
            pdf.cell(CW[0], 8, str(stop.get('seq', idx + 1)), border=1, fill=even, align='C')
            pdf.cell(CW[1], 8, cust_s, border=1, fill=even)
            pdf.cell(CW[2], 8, addr_s, border=1, fill=even)
            pdf.cell(CW[3], 8, vol_s,  border=1, fill=even, align='R')
            pdf.set_font('Helvetica', 'U', 8)
            pdf.set_text_color(0, 80, 200)
            pdf.cell(CW[4], 8, 'Open Map', border=1, align='C', link=maps_url, fill=even)
            pdf.set_text_color(0, 0, 0)
            pdf.ln()

        # ── Totals bar ─────────────────────────────────────
        pdf.ln(2)
        pdf.set_fill_color(232, 240, 254)
        pdf.set_font('Helvetica', 'B', 8.5)
        pdf.cell(0, 8,
            f"  TOTAL: {truck.get('acc_weight',0):,.1f} kg   |   "
            f"{stops_n} drops   |   {util_pct}% utilization   |   "
            f"Est. {truck.get('total_km',0)} km",
            fill=True, ln=True, border=1)

        # ── Signature block ────────────────────────────────
        pdf.ln(5)
        pdf.set_font('Helvetica', '', 8.5)
        for label in ['Driver / Trucker:', 'Plate #:', 'Dispatcher:', 'Date Delivered:']:
            pdf.cell(40, 7, label, border='B')
            pdf.cell(8, 7, '')
        pdf.ln(8)

        # ── Output ─────────────────────────────────────────
        out   = io.BytesIO(bytes(pdf.output()))
        fname = f"{truck_id}_route_sheet.pdf"
        return send_file(out, mimetype='application/pdf',
                         as_attachment=True, download_name=fname)

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/api/save-plan', methods=['POST'])
def save_plan():
    try:
        router    = request.args.get('router', 'default')
        temp_plan = get_router_temp('plan', router)
        if not os.path.exists(temp_plan):
            return jsonify({'error': 'No plan found — please generate a route plan first'}), 400
        body       = request.get_json() or {}
        plan_date  = body.get('date', datetime.now().strftime('%Y-%m-%d'))
        plan_name  = body.get('name', f'Route Plan {plan_date} {datetime.now().strftime("%H:%M")}')
        created_by = body.get('created_by', router)
        with open(temp_plan) as fp:
            trucks = json.load(fp)
        if not trucks:
            return jsonify({'error': 'Plan is empty — no trucks to save'}), 400

        # ── Collaborative merge ──────────────────────────────────────────
        # Before saving, pull in trucks from other planners for the same
        # date so that every Save produces the full team plan, not just
        # this planner's slice.
        conn_m = open_db(); cm = conn_m.cursor()
        cm.execute('''SELECT created_by, data FROM route_plans
                      WHERE plan_date=? AND created_by != ?
                      ORDER BY created_at ASC''', (plan_date, created_by))
        other_rows = cm.fetchall(); conn_m.close()

        my_ids = {t.get('truck_id', '') for t in trucks}
        for _, other_data in other_rows:
            try:
                for ot in json.loads(other_data):
                    tid = ot.get('truck_id', '')
                    if tid and tid not in my_ids:
                        trucks.append(ot)
                        my_ids.add(tid)
            except Exception:
                pass
        # ────────────────────────────────────────────────────────────────

        def safe_float(v):
            try: return float(v or 0)
            except Exception: return 0.0
        summary = {
            'truck_count':     len(trucks),
            'total_drops':     sum(len(t['stops']) for t in trucks),
            'total_volume':    round(sum(safe_float(t.get('acc_weight')) for t in trucks), 2),
            'avg_utilization': round((sum(safe_float(t.get('util_pct')) for t in trucks)/len(trucks))*100,1) if trucks else 0,
            'total_km':        round(sum(safe_float(t.get('total_km')) for t in trucks), 1),
            'total_cost':      round(sum(safe_float(t.get('truck_rate')) for t in trucks), 2),
        }
        conn = open_db()
        c = conn.cursor()
        # If THIS user already saved a plan for this date, update it — never overwrite another user's plan
        c.execute('SELECT id FROM route_plans WHERE plan_date=? AND created_by=? ORDER BY id DESC LIMIT 1', (plan_date, created_by))
        existing = c.fetchone()
        if existing:
            plan_id = existing[0]
            c.execute('''UPDATE route_plans
                         SET plan_name=?, created_by=?, data=?, summary=?, status=?, created_at=datetime('now','localtime')
                         WHERE id=?''',
                      (plan_name, created_by, json.dumps(trucks, default=str), json.dumps(summary), 'finalized', plan_id))
        else:
            c.execute('INSERT INTO route_plans (plan_date,plan_name,created_by,data,summary,status) VALUES (?,?,?,?,?,?)',
                      (plan_date, plan_name, created_by, json.dumps(trucks, default=str), json.dumps(summary), 'finalized'))
            plan_id = c.lastrowid
        conn.commit(); conn.close()
        # Auto-upload to Google Drive
        drive_url = None
        try:
            drive_url = _gdrive_upload_plan_excel(trucks, plan_date, plan_name)
        except Exception:
            pass  # Drive upload is best-effort; don't fail the save
        resp = {'success': True, 'plan_id': plan_id, 'plan_name': plan_name}
        if drive_url:
            resp['drive_url'] = drive_url
        return jsonify(resp)
    except Exception as e:
        import traceback
        return jsonify({'error': f'DB save error: {str(e)}', 'trace': traceback.format_exc()}), 500


@app.route('/api/history', methods=['GET'])
def get_history():
    try:
        conn = open_db()
        c = conn.cursor()
        c.execute('SELECT id,plan_date,plan_name,created_at,created_by,status,summary FROM route_plans ORDER BY created_at DESC LIMIT 60')
        rows = c.fetchall()
        conn.close()
        plans = []
        for row in rows:
            try: summary = json.loads(row[6]) if row[6] else {}
            except Exception: summary = {}
            plans.append({'id':row[0],'date':row[1],'name':row[2],'created_at':row[3],
                          'created_by':row[4] or '','status':row[5],'summary':summary})
        return jsonify({'plans': plans})
    except Exception as e:
        return jsonify({'plans': [], 'error': str(e)})


@app.route('/api/latest-plan-meta', methods=['GET'])
def latest_plan_meta():
    """Lightweight poll endpoint — returns just the ID + metadata of the newest saved plan."""
    conn = open_db(); c = conn.cursor()
    c.execute('SELECT id, plan_date, plan_name, created_at, created_by, summary '
              'FROM route_plans ORDER BY id DESC LIMIT 1')
    row = c.fetchone(); conn.close()
    if not row:
        return jsonify({'plan_id': None})
    try:    summary = json.loads(row[5]) if row[5] else {}
    except: summary = {}
    return jsonify({
        'plan_id':    row[0],
        'plan_date':  row[1],
        'plan_name':  row[2],
        'created_at': row[3],
        'created_by': row[4] or '',
        'truck_count':  summary.get('truck_count', 0),
        'total_drops':  summary.get('total_drops', 0),
    })


@app.route('/api/date-plans/<date>', methods=['GET'])
def get_date_plans(date):
    """Return all trucks from ALL planners' saved plans for a given date, merged by truck_id.
    Used by the frontend to do collaborative merge-on-load rather than destructive replace."""
    try:
        conn = open_db(); c = conn.cursor()
        c.execute('''SELECT created_by, data, created_at
                     FROM route_plans WHERE plan_date=?
                     ORDER BY created_at ASC''', (date,))
        rows = c.fetchall(); conn.close()
        # Merge: later saves for the same truck_id win (most recent edit wins)
        merged = {}
        planners = set()
        for created_by, data_json, _ in rows:
            try:
                trucks = json.loads(data_json)
                planners.add(created_by or 'unknown')
                for t in trucks:
                    tid = t.get('truck_id', '')
                    if tid:
                        merged[tid] = dict(t, _saved_by=created_by or '')
            except Exception:
                pass
        return jsonify({'trucks': list(merged.values()), 'planners': sorted(planners), 'date': date})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/history/<int:plan_id>', methods=['GET'])
def get_history_plan(plan_id):
    router    = request.args.get('router', 'default')
    temp_plan = get_router_temp('plan', router)
    conn = open_db()
    c = conn.cursor()
    c.execute('SELECT id,plan_date,plan_name,created_at,created_by,status,data,summary FROM route_plans WHERE id=?',(plan_id,))
    row = c.fetchone(); conn.close()
    if not row: return jsonify({'error': 'Plan not found'}), 404
    trucks = json.loads(row[6])
    with open(temp_plan, 'w') as fp:
        json.dump(trucks, fp)
    return jsonify({'id':row[0],'date':row[1],'name':row[2],'created_at':row[3],
                    'created_by':row[4] or '','status':row[5],'trucks':trucks,
                    'summary':json.loads(row[7]) if row[7] else {}})


@app.route('/api/history/<int:plan_id>', methods=['DELETE'])
def delete_history_plan(plan_id):
    try:
        conn = open_db(); c = conn.cursor()
        c.execute('SELECT id FROM route_plans WHERE id=?', (plan_id,))
        if not c.fetchone():
            conn.close()
            return jsonify({'error': 'Plan not found'}), 404
        c.execute('DELETE FROM route_plans WHERE id=?', (plan_id,))
        conn.commit(); conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/monitoring-history', methods=['GET'])
def get_monitoring_history():
    """Return summary of all saved monitoring dates (newest first)."""
    try:
        conn = open_db(); c = conn.cursor()
        c.execute("""
            SELECT plan_date,
                   COUNT(*) as total_stops,
                   SUM(CASE WHEN status='DONE' THEN 1 ELSE 0 END) as done_count,
                   SUM(CASE WHEN otif_status LIKE '%OTIF%' THEN 1 ELSE 0 END) as otif_count,
                   COUNT(DISTINCT truck_id) as truck_count
            FROM monitoring_records
            GROUP BY plan_date
            ORDER BY plan_date DESC
            LIMIT 90
        """)
        rows = c.fetchall(); conn.close()
        result = []
        for row in rows:
            plan_date, total, done, otif, trucks = row
            result.append({
                'plan_date': plan_date,
                'total_stops': total or 0,
                'done_count': done or 0,
                'otif_count': otif or 0,
                'truck_count': trucks or 0,
                'completion_pct': round((done or 0) / total * 100) if total else 0,
                'otif_pct': round((otif or 0) / total * 100) if total else 0,
            })
        return jsonify({'dates': result})
    except Exception as e:
        return jsonify({'dates': [], 'error': str(e)})


@app.route('/api/monitoring-history/<plan_date>', methods=['GET'])
def get_monitoring_history_detail(plan_date):
    """Return full monitoring records for a specific date."""
    try:
        conn = open_db(); c = conn.cursor()
        c.execute('''SELECT truck_id,truck_type,seq,customer_name,cluster_id,area,
                            status,receiving_time,actual_done,otif_status,concerns,remarks,trucker_code
                     FROM monitoring_records WHERE plan_date=?
                     ORDER BY truck_id,seq''', (plan_date,))
        rows = c.fetchall(); conn.close()
        cols = ['truck_id','truck_type','seq','customer_name','cluster_id','area',
                'status','receiving_time','actual_done','otif_status','concerns','remarks','trucker_code']
        return jsonify({'records': [dict(zip(cols, r)) for r in rows]})
    except Exception as e:
        return jsonify({'records': [], 'error': str(e)})


@app.route('/api/monitoring-db', methods=['DELETE'])
def delete_monitoring_db():
    """Delete all monitoring records for a given date."""
    try:
        date = request.args.get('date', '')
        if not date:
            return jsonify({'error': 'date param required'}), 400
        conn = open_db(); c = conn.cursor()
        c.execute('DELETE FROM monitoring_records WHERE plan_date=?', (date,))
        deleted = c.rowcount
        conn.commit(); conn.close()
        return jsonify({'success': True, 'deleted': deleted})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/monitoring', methods=['GET'])
def get_monitoring():
    router    = request.args.get('router', 'default')
    temp_plan = get_router_temp('plan', router)
    temp_mon  = get_router_temp('monitoring', router)
    if not os.path.exists(temp_plan):
        return jsonify({'error': 'No plan loaded'}), 400
    with open(temp_plan) as fp: trucks = json.load(fp)
    existing = {}
    if os.path.exists(temp_mon):
        with open(temp_mon) as fp:
            for m in json.load(fp): existing[m['truck_id']+'|'+str(m['seq'])] = m
    rows = []
    for truck in trucks:
        total = len(truck['stops'])
        for stop in truck['stops']:
            k = truck['truck_id']+'|'+str(stop['seq'])
            saved = existing.get(k, {})
            rows.append({'truck_id':truck['truck_id'],'truck_type':truck['truck_type'],
                         'seq':stop['seq'],'total_drops':total,'customer_name':stop['customer_name'],
                         'cluster_id':stop['cluster_id'],'area':stop.get('match_area',''),
                         'receiving_time':saved.get('receiving_time','06:00 PM'),
                         'status':saved.get('status','PENDING'),'actual_done':saved.get('actual_done',''),
                         'otif_status':saved.get('otif_status',''),'concerns':saved.get('concerns','NONE'),
                         'remarks':saved.get('remarks',''),'trucker_code':truck.get('trucker_code',''),
                         'shipping_address':stop.get('shipping_address','')})
    return jsonify({'monitoring': rows})


@app.route('/api/update-monitoring', methods=['POST'])
def update_monitoring():
    router   = request.args.get('router', 'default')
    temp_mon = get_router_temp('monitoring', router)
    data     = request.get_json()
    if not data or 'monitoring' not in data:
        return jsonify({'error': 'Invalid payload'}), 400
    rows = data['monitoring']
    for row in rows:
        if row.get('status')=='DONE' and row.get('actual_done') and not row.get('otif_status'):
            try:
                def parse_time(s):
                    s=str(s).upper().strip(); pm='PM' in s
                    nums=re.sub(r'[APM\s]','',s).split(':')
                    h,m=int(nums[0]),int(nums[1]) if len(nums)>1 else 0
                    if pm and h!=12: h+=12
                    if not pm and h==12: h=0
                    return h*60+m
                sched=parse_time(row.get('receiving_time','18:00'))
                actual=parse_time(row['actual_done'])
                row['otif_status']='✅ OTIF' if actual<=sched else '❌ LATE'
            except Exception: pass
    with open(temp_mon, 'w') as fp: json.dump(rows, fp)
    return jsonify({'success': True})


@app.route('/api/save-monitoring-db', methods=['POST'])
def save_monitoring_db():
    try:
        data=request.get_json() or {}
        monitoring=data.get('monitoring',[]); plan_id=data.get('plan_id')
        plan_date=data.get('plan_date',datetime.now().strftime('%Y-%m-%d'))
        if not monitoring: return jsonify({'error':'No monitoring data to save'}), 400
        conn=open_db(); c=conn.cursor()
        c.execute('DELETE FROM monitoring_records WHERE plan_date=?',(plan_date,))
        saved = 0
        for row in monitoring:
            c.execute('''INSERT INTO monitoring_records
                (plan_id,plan_date,truck_id,truck_type,seq,customer_name,cluster_id,area,
                 status,receiving_time,actual_done,otif_status,concerns,remarks,trucker_code,
                 shipping_address,food_safety_issue,food_safety_detail,crew_issue,crew_detail,
                 do_number,return_date,pod_remarks)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                      (plan_id,plan_date,row.get('truck_id'),row.get('truck_type'),row.get('seq'),
                       row.get('customer_name'),row.get('cluster_id'),row.get('area'),row.get('status'),
                       row.get('receiving_time'),row.get('actual_done'),row.get('otif_status'),
                       row.get('concerns'),row.get('remarks'),row.get('trucker_code'),
                       row.get('shipping_address',''),
                       row.get('food_safety_issue','NONE'), row.get('food_safety_detail',''),
                       row.get('crew_issue','NONE'),        row.get('crew_detail',''),
                       row.get('do_number',''),             row.get('return_date',''),
                       row.get('pod_remarks','')))
            saved += 1
        conn.commit(); conn.close()
        return jsonify({'success': True, 'saved': saved})
    except Exception as e:
        import traceback
        return jsonify({'error': f'Monitoring DB save error: {str(e)}', 'trace': traceback.format_exc()}), 500


# ─────────────────────────────────────────────────────────────
#  COLD CHAIN COMPLIANCE  (reads from monitoring_records)
# ─────────────────────────────────────────────────────────────
@app.route('/api/cold-chain', methods=['GET'])
def get_cold_chain():
    """
    GET /api/cold-chain          → list of dates that have monitoring records
    GET /api/cold-chain?date=X   → all monitoring rows for that date (cold chain view)
    """
    try:
        date = request.args.get('date', '')
        conn = open_db(); c = conn.cursor()
        if date:
            c.execute('''SELECT id, plan_date, truck_id, truck_type, customer_name, do_number,
                                trucker_code, food_safety_issue, food_safety_detail,
                                crew_issue, crew_detail
                         FROM monitoring_records
                         WHERE plan_date=? ORDER BY truck_id, seq''', (date,))
            cols = [d[0] for d in c.description]
            rows = [dict(zip(cols, r)) for r in c.fetchall()]
            conn.close()
            return jsonify({'records': rows})
        else:
            c.execute('''SELECT DISTINCT plan_date FROM monitoring_records
                         WHERE plan_date IS NOT NULL AND plan_date != ''
                         ORDER BY plan_date DESC''')
            dates = [r[0] for r in c.fetchall()]
            conn.close()
            return jsonify({'dates': dates})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def _period_group_expr(period):
    """Return SQL expression to group plan_date by period."""
    if period == 'weekly':
        return "strftime('%Y-W%W', plan_date)"
    if period == 'monthly':
        return "strftime('%Y-%m', plan_date)"
    if period == 'quarterly':
        return "(strftime('%Y', plan_date) || '-Q' || ((CAST(strftime('%m', plan_date) AS INTEGER) - 1) / 3 + 1))"
    if period == 'yearly':
        return "strftime('%Y', plan_date)"
    return 'plan_date'  # daily (default)

@app.route('/api/cold-chain/summary', methods=['GET'])
def cold_chain_summary():
    """Returns compliance % grouped by period, sourced from monitoring_records."""
    try:
        period = request.args.get('period', 'daily')
        grp = _period_group_expr(period)
        conn = open_db(); c = conn.cursor()
        c.execute(f'''SELECT {grp} as period_label,
                            MIN(plan_date) as date,
                            COUNT(*) as total,
                            SUM(CASE WHEN food_safety_issue='NONE' OR food_safety_issue IS NULL THEN 1 ELSE 0 END) as ok,
                            SUM(CASE WHEN food_safety_issue IS NOT NULL AND food_safety_issue!='NONE' THEN 1 ELSE 0 END) as fs_issues,
                            SUM(CASE WHEN crew_issue='NONE' OR crew_issue IS NULL THEN 1 ELSE 0 END) as crew_ok,
                            SUM(CASE WHEN crew_issue IS NOT NULL AND crew_issue!='NONE' THEN 1 ELSE 0 END) as crew_issues
                     FROM monitoring_records
                     WHERE plan_date IS NOT NULL AND plan_date != ''
                     GROUP BY {grp} ORDER BY {grp} DESC LIMIT 30''')
        rows = []
        for r in c.fetchall():
            total = r[2] or 1
            rows.append({
                'label': r[0], 'date': r[1], 'total': r[2],
                'ok': r[3], 'issues': r[4],
                'compliance_pct': round(r[3]/total*100, 1),
                'crew_ok': r[5], 'crew_issues': r[6],
                'crew_pct': round(r[5]/total*100, 1)
            })
        conn.close()
        return jsonify({'summary': rows})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ─────────────────────────────────────────────────────────────
#  POD RETURN COMPLIANCE  (reads/updates monitoring_records)
# ─────────────────────────────────────────────────────────────
@app.route('/api/pod', methods=['GET'])
def get_pod():
    """
    GET /api/pod              → list of dates with monitoring records
    GET /api/pod?date=X       → monitoring rows for that date with POD fields
    """
    try:
        date = request.args.get('date', '')
        conn = open_db(); c = conn.cursor()
        if date:
            c.execute('''SELECT id, plan_date, truck_id, customer_name, do_number,
                                area, status, return_date, pod_remarks
                         FROM monitoring_records
                         WHERE plan_date=? ORDER BY truck_id, seq''', (date,))
            cols = [d[0] for d in c.description]
            rows = []
            for r in c.fetchall():
                row = dict(zip(cols, r))
                # compute aging
                row['days_aging'] = None
                row['pod_status'] = 'PENDING'
                if row.get('return_date'):
                    try:
                        d1 = datetime.strptime(row['plan_date'], '%Y-%m-%d')
                        d2 = datetime.strptime(row['return_date'], '%Y-%m-%d')
                        row['days_aging'] = (d2 - d1).days
                        row['pod_status'] = 'ON-TIME' if row['days_aging'] <= 2 else 'LATE'
                    except Exception:
                        pass
                rows.append(row)
            conn.close()
            return jsonify({'records': rows})
        else:
            c.execute('''SELECT DISTINCT plan_date FROM monitoring_records
                         WHERE plan_date IS NOT NULL AND plan_date != ''
                         ORDER BY plan_date DESC''')
            dates = [r[0] for r in c.fetchall()]
            conn.close()
            return jsonify({'dates': dates})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pod/summary', methods=['GET'])
def pod_summary():
    """Returns POD compliance % grouped by period for historical trend tiles."""
    try:
        period = request.args.get('period', 'daily')
        grp = _period_group_expr(period)
        conn = open_db(); c = conn.cursor()
        c.execute(f'''SELECT {grp} as period_label,
                            MIN(plan_date) as date,
                            COUNT(*) as total,
                            SUM(CASE WHEN return_date IS NOT NULL AND return_date!='' THEN 1 ELSE 0 END) as returned,
                            SUM(CASE WHEN return_date IS NOT NULL AND return_date!=''
                                      AND (julianday(return_date)-julianday(plan_date))<=2 THEN 1 ELSE 0 END) as ontime
                     FROM monitoring_records
                     WHERE plan_date IS NOT NULL AND plan_date!=''
                     GROUP BY {grp} ORDER BY {grp} DESC LIMIT 30''')
        rows = []
        for r in c.fetchall():
            ret = r[3] or 0
            ontime = r[4] or 0
            pct = round(ontime/ret*100, 1) if ret else 0
            rows.append({'label':r[0],'date':r[1],'total':r[2],'returned':ret,'ontime':ontime,'compliance_pct':pct})
        conn.close()
        return jsonify({'summary': rows})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pod', methods=['POST'])
def save_pod():
    """Update return_date and pod_remarks on individual monitoring_records rows."""
    try:
        data = request.get_json() or {}
        records = data.get('records', [])
        conn = open_db(); c = conn.cursor()
        saved = 0
        for r in records:
            if not r.get('id'):
                continue
            c.execute('''UPDATE monitoring_records
                         SET return_date=?, pod_remarks=?,
                             saved_at=datetime('now','localtime')
                         WHERE id=?''',
                      (r.get('return_date', ''), r.get('pod_remarks', ''), r['id']))
            saved += 1
        conn.commit(); conn.close()
        return jsonify({'success': True, 'saved': saved})
    except Exception as e:
        return jsonify({'error': str(e)}), 500




@app.route('/api/analytics', methods=['GET'])
def get_analytics():
    """
    GET /api/analytics?period=daily|weekly|monthly|quarterly|yearly
    Returns aggregated performance metrics.
    Primary source: route_plans (always has data when a plan is saved).
    Secondary (overlaid): monitoring_records (for Done%, OTIF%, issues).
    """
    try:
        period = request.args.get('period', 'daily')

        def period_expr(col, p):
            if p == 'weekly':    return f"strftime('%Y-W%W', {col})"
            if p == 'monthly':   return f"strftime('%Y-%m', {col})"
            if p == 'quarterly': return f"(strftime('%Y', {col}) || '-Q' || ((CAST(strftime('%m', {col}) AS INTEGER) - 1) / 3 + 1))"
            if p == 'yearly':    return f"strftime('%Y', {col})"
            return col

        grp = period_expr('plan_date', period)
        conn = open_db(); c = conn.cursor()

        # PRIMARY: aggregate from route_plans (every saved plan shows up here)
        c.execute(f"""
            SELECT
                {grp}                                                               AS period_key,
                MIN(plan_date)                                                      AS sample_date,
                SUM(CAST(json_extract(summary, '$.truck_count') AS REAL))           AS plan_trucks,
                SUM(CAST(json_extract(summary, '$.total_drops') AS REAL))           AS plan_drops,
                SUM(CAST(json_extract(summary, '$.total_volume') AS REAL))          AS total_vol,
                SUM(CAST(json_extract(summary, '$.total_cost')   AS REAL))          AS total_cost
            FROM route_plans
            WHERE plan_date IS NOT NULL AND summary IS NOT NULL AND summary != ''
            GROUP BY {grp}
            ORDER BY {grp} DESC
            LIMIT 60
        """)
        plan_rows = c.fetchall()
        plan_cols = [d[0] for d in c.description]

        # SECONDARY: aggregate from monitoring_records (overlaid when available)
        grp_mon = period_expr('plan_date', period)
        c.execute(f"""
            SELECT
                {grp_mon}                                                                   AS period_key,
                COUNT(*)                                                                    AS total_drops,
                COUNT(DISTINCT truck_id)                                                    AS trucks,
                SUM(CASE WHEN status='DONE' THEN 1 ELSE 0 END)                             AS done_count,
                SUM(CASE WHEN otif_status LIKE '%OTIF%' THEN 1 ELSE 0 END)                 AS otif_count,
                SUM(CASE WHEN food_safety_issue != 'NONE' AND food_safety_issue IS NOT NULL THEN 1 ELSE 0 END) AS fs_issues,
                SUM(CASE WHEN crew_issue != 'NONE' AND crew_issue IS NOT NULL THEN 1 ELSE 0 END) AS crew_issues,
                SUM(CASE WHEN concerns IS NOT NULL AND concerns != '' AND concerns != 'NONE' THEN 1 ELSE 0 END) AS concerns_count,
                SUM(CASE WHEN return_date IS NOT NULL AND return_date != '' THEN 1 ELSE 0 END) AS pod_returned
            FROM monitoring_records
            WHERE plan_date IS NOT NULL AND plan_date != ''
            GROUP BY {grp_mon}
        """)
        mon_map = {}
        for r in c.fetchall():
            mon_map[r[0]] = {
                'total_drops':    r[1] or 0,
                'trucks':         r[2] or 0,
                'done_count':     r[3] or 0,
                'otif_count':     r[4] or 0,
                'fs_issues':      r[5] or 0,
                'crew_issues':    r[6] or 0,
                'concerns_count': r[7] or 0,
                'pod_returned':   r[8] or 0,
            }
        conn.close()

        period_labels = {
            'daily': 'Date', 'weekly': 'Week', 'monthly': 'Month',
            'quarterly': 'Quarter', 'yearly': 'Year'
        }

        result = []
        for row in plan_rows:
            d = dict(zip(plan_cols, row))
            pk = d['period_key']
            mon = mon_map.get(pk, {})

            total_vol  = d['total_vol']  or 0
            total_cost = d['total_cost'] or 0
            plan_trucks = int(d['plan_trucks'] or 0)
            plan_drops  = int(d['plan_drops']  or 0)

            # Use monitoring counts when available, fall back to plan summary
            drops  = mon.get('total_drops') or plan_drops
            trucks = mon.get('trucks')       or plan_trucks
            done   = mon.get('done_count',     0)
            otif   = mon.get('otif_count',     0)
            fs     = mon.get('fs_issues',      0)
            crew   = mon.get('crew_issues',    0)
            concerns  = mon.get('concerns_count', 0)
            pod_ret   = mon.get('pod_returned',   0)

            cpk = round(total_cost / total_vol, 2) if total_vol > 0 else 0
            result.append({
                'period':       pk,
                'label':        pk,
                'sample_date':  d['sample_date'],
                'trucks':       trucks,
                'drops':        drops,
                'total_vol':    round(total_vol, 1),
                'done_count':   done,
                'done_pct':     round(done / drops * 100, 1) if drops else 0,
                'otif_count':   otif,
                'otif_pct':     round(otif / drops * 100, 1) if drops else 0,
                'fs_issues':    fs,
                'crew_issues':  crew,
                'concerns':     concerns,
                'pod_returned': pod_ret,
                'total_cost':   round(total_cost, 2),
                'cost_per_kg':  cpk,
            })

        return jsonify({'rows': result, 'period': period, 'period_label': period_labels.get(period, 'Period')})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/debug/db', methods=['GET'])
def debug_db():
    """Debug endpoint — shows row counts and sample data from key tables."""
    try:
        conn = open_db(); c = conn.cursor()
        c.execute('SELECT COUNT(*) FROM monitoring_records'); mon_count = c.fetchone()[0]
        c.execute('SELECT COUNT(*) FROM route_plans');        plan_count = c.fetchone()[0]
        c.execute('SELECT plan_date, COUNT(*) FROM monitoring_records GROUP BY plan_date ORDER BY plan_date DESC LIMIT 10')
        mon_dates = [{'date': r[0], 'count': r[1]} for r in c.fetchall()]
        c.execute('SELECT plan_date, plan_name, created_by FROM route_plans ORDER BY id DESC LIMIT 5')
        plans = [{'date': r[0], 'name': r[1], 'by': r[2]} for r in c.fetchall()]
        conn.close()
        return jsonify({
            'db_path': DB_PATH,
            'monitoring_records_total': mon_count,
            'route_plans_total': plan_count,
            'monitoring_by_date': mon_dates,
            'recent_plans': plans,
        })
    except Exception as e:
        return jsonify({'error': str(e), 'db_path': DB_PATH}), 500


@app.route('/api/rates', methods=['GET'])
def get_rates():
    trucker = request.args.get('trucker', '')
    wh      = request.args.get('wh', '')
    search  = request.args.get('search', '').strip()
    try:
        conn = open_db(); c = conn.cursor()
        q = 'SELECT id,trucker,capacity_kg,pickup_wh,area,rate,cost_per_kg,is_active,updated_at FROM rate_master WHERE 1=1'
        params = []
        if trucker: q += ' AND UPPER(trucker)=?'; params.append(trucker.upper())
        if wh:      q += ' AND UPPER(pickup_wh)=?'; params.append(wh.upper())
        if search:
            q += ' AND (UPPER(trucker) LIKE ? OR UPPER(area) LIKE ? OR UPPER(pickup_wh) LIKE ?)'
            s = '%' + search.upper() + '%'
            params += [s, s, s]
        q += ' ORDER BY trucker,capacity_kg,pickup_wh,area'
        c.execute(q, params); rows = c.fetchall(); conn.close()
        return jsonify({'rates': [{'id':r[0],'trucker':r[1],'capacity_kg':r[2],'pickup_wh':r[3],
                                   'area':r[4],'rate':r[5],'cost_per_kg':r[6],'is_active':r[7],'updated_at':r[8]}
                                  for r in rows]})
    except Exception as e:
        return jsonify({'rates': [], 'error': str(e)})


@app.route('/api/rates/import', methods=['POST'])
def import_rates():
    if 'file' not in request.files: return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']; filename = secure_filename(f.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'rates_' + filename); f.save(filepath)
    try:
        pd = _get_pd()
        # Try with header first
        df = pd.read_excel(filepath, header=0)
        cols_norm = [str(c).strip().lower().replace(' ', '_') for c in df.columns]

        # Column name aliases — handles "TRUCKING", "trucker", "PICK UP LOCATION", etc.
        col_map_aliases = {
            'trucking': 'trucker', 'trucker': 'trucker',
            'std_capacity': 'capacity_kg', 'capacity_kg': 'capacity_kg', 'capacity': 'capacity_kg',
            'pick_up_location': 'pickup_wh', 'pickup_wh': 'pickup_wh', 'wh': 'pickup_wh',
            'area': 'area',
            'rate': 'rate',
            'cost_per_kg': 'cost_per_kg', 'cpk': 'cost_per_kg',
        }
        mapped = {col_map_aliases.get(c, c): orig for c, orig in zip(cols_norm, df.columns)}

        if 'trucker' not in mapped:
            # No recognisable headers — treat as headerless, use positional mapping
            df = pd.read_excel(filepath, header=None)
            if len(df.columns) >= 5:
                df = df.iloc[:, :6] if len(df.columns) >= 6 else df
                pos_cols = ['trucker', 'capacity_kg', 'pickup_wh', 'area', 'rate', 'cost_per_kg']
                df.columns = pos_cols[:len(df.columns)]
            else:
                return jsonify({'error': 'File must have at least 5 columns: Trucker, Capacity, WH, Area, Rate'}), 400
        else:
            # Rename to standard names
            df = df.rename(columns={v: k for k, v in mapped.items() if k in col_map_aliases.values()})
            df.columns = [col_map_aliases.get(str(c).strip().lower().replace(' ', '_'), c) for c in df.columns]
    except Exception as e:
        return jsonify({'error': str(e)}), 400

    df = df[df['trucker'].astype(str).str.strip().str.len() > 0]
    df = df[df['trucker'].astype(str) != 'nan']

    def safe_float(v):
        try: return float(str(v).replace('₱', '').replace(',', '').strip() or 0)
        except: return 0.0

    conn = open_db(); c = conn.cursor()
    c.execute('DELETE FROM rate_master')
    inserted = 0
    for _, row in df.iterrows():
        try:
            trucker   = str(row.get('trucker', '')).strip()
            cap       = int(float(row.get('capacity_kg', 0) or 0))
            wh        = str(row.get('pickup_wh', '')).strip().upper()
            area      = str(row.get('area', '')).strip().upper()
            rate      = safe_float(row.get('rate', 0))
            cpk_raw   = row.get('cost_per_kg', None)
            cpk       = safe_float(cpk_raw) if cpk_raw and str(cpk_raw) not in ('nan', 'None', '') else (rate / cap if cap > 0 else 0)
            if not trucker or trucker == 'nan': continue
            c.execute('INSERT INTO rate_master (trucker,capacity_kg,pickup_wh,area,rate,cost_per_kg) VALUES (?,?,?,?,?,?)',
                      (trucker, cap, wh, area, rate, round(cpk, 4)))
            inserted += 1
        except Exception:
            pass
    conn.commit(); conn.close()
    return jsonify({'success': True, 'imported': inserted})


@app.route('/api/rates/<int:rate_id>', methods=['PUT'])
def update_rate(rate_id):
    data=request.get_json() or {}
    conn=open_db(); c=conn.cursor()
    c.execute("UPDATE rate_master SET rate=?,cost_per_kg=?,is_active=?,updated_at=datetime('now','localtime') WHERE id=?",
              (data.get('rate',0),data.get('cost_per_kg',0),1 if data.get('is_active',True) else 0,rate_id))
    conn.commit(); conn.close(); return jsonify({'success':True})


@app.route('/api/rates/<int:rate_id>', methods=['DELETE'])
def delete_rate(rate_id):
    conn=open_db(); c=conn.cursor()
    c.execute('UPDATE rate_master SET is_active=0 WHERE id=?',(rate_id,))
    conn.commit(); conn.close(); return jsonify({'success':True})


@app.route('/api/rates/add', methods=['POST'])
def add_rate():
    data=request.get_json() or {}
    conn=open_db(); c=conn.cursor()
    c.execute('INSERT INTO rate_master (trucker,capacity_kg,pickup_wh,area,rate,cost_per_kg) VALUES (?,?,?,?,?,?)',
              (str(data.get('trucker','')).strip(),int(data.get('capacity_kg',0)),
               str(data.get('pickup_wh','')).upper(),str(data.get('area','')).upper(),
               float(data.get('rate',0)),float(data.get('cost_per_kg',0))))
    new_id=c.lastrowid; conn.commit(); conn.close()
    return jsonify({'success':True,'id':new_id})


@app.route('/api/plan-cost', methods=['GET'])
def get_plan_cost():
    router    = request.args.get('router', 'default')
    temp_plan = get_router_temp('plan', router)
    if not os.path.exists(temp_plan): return jsonify({'costs': []})
    with open(temp_plan) as fp: trucks = json.load(fp)
    costs = []
    for truck in trucks:
        trucker   = truck.get('trucker_code', '').strip()
        all_areas = [s.get('match_area', '') for s in truck.get('stops', [])]
        # Use highest rate across all areas (farthest/most expensive destination pays for the trip)
        rate, cpk = lookup_highest_rate(
            trucker, truck.get('truck_cap', 0), truck.get('wh_group', ''), all_areas
        )
        # Fallback: find the area with highest volume for display
        area_vols = {}
        for s in truck.get('stops', []):
            a = s.get('match_area', '')
            area_vols[a] = area_vols.get(a, 0) + s.get('vol', 0)
        primary_area = max(area_vols, key=area_vols.get) if area_vols else ''
        costs.append({
            'truck_id':     truck['truck_id'],
            'rate':         rate,
            'cost_per_kg':  cpk,
            'primary_area': primary_area,
            'all_areas':    list(set(all_areas)),
        })
    return jsonify({'costs': costs})


@app.route('/api/truck-rate', methods=['GET'])
def get_truck_rate():
    """Look up rate for a single truck — called when trucker code is entered."""
    trucker  = request.args.get('trucker', '').strip()
    cap      = int(request.args.get('cap', 0) or 0)
    wh_group = request.args.get('wh_group', '')
    areas_raw = request.args.get('areas', '')
    areas    = [a.strip() for a in areas_raw.split(',') if a.strip()] if areas_raw else []

    if not trucker:
        return jsonify({'rate': None, 'cost_per_kg': None, 'best_area': ''})

    rate, cpk = lookup_highest_rate(trucker, cap, wh_group, areas)

    # Find which area gave the highest rate (multi-candidate lookup)
    best_area = ''
    if areas and rate is not None:
        rate_wh    = WH_TO_RATE_WH.get(wh_group, wh_group).upper()
        trucker_up = trucker.upper()
        try:
            conn = open_db()
            c    = conn.cursor()
            best_r = None
            for area in set(areas):
                for candidate in area_candidates(str(area)):
                    c.execute('''SELECT rate, area FROM rate_master
                                 WHERE UPPER(trucker)=? AND capacity_kg=?
                                   AND UPPER(pickup_wh)=? AND UPPER(area)=?
                                   AND is_active=1 LIMIT 1''',
                              (trucker_up, cap, rate_wh, candidate.upper()))
                    row = c.fetchone()
                    if row and (best_r is None or float(row[0]) > best_r):
                        best_r    = float(row[0])
                        best_area = row[1]
                    if row:
                        break  # found a match for this area, stop trying candidates
            conn.close()
        except Exception:
            pass

    return jsonify({'rate': rate, 'cost_per_kg': cpk, 'best_area': best_area})


@app.route('/api/rate-debug', methods=['GET'])
def rate_debug():
    """Return all rate master entries for a given trucker (for diagnosis)."""
    trucker = request.args.get('trucker', '').strip().upper()
    if not trucker:
        return jsonify({'error': 'trucker param required'}), 400
    try:
        conn = open_db()
        c    = conn.cursor()
        c.execute('''SELECT id, trucker, capacity_kg, pickup_wh, area, rate, cost_per_kg, is_active
                     FROM rate_master WHERE UPPER(trucker)=? ORDER BY pickup_wh, capacity_kg, area''',
                  (trucker,))
        rows = [{'id':r[0],'trucker':r[1],'capacity_kg':r[2],'pickup_wh':r[3],
                 'area':r[4],'rate':r[5],'cost_per_kg':r[6],'is_active':r[7]} for r in c.fetchall()]
        conn.close()
        return jsonify({'trucker': trucker, 'entries': rows, 'count': len(rows)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ─────────────────────────────────────────────────────────────
#  BATCH ROUTING — merge new orders into existing plan
# ─────────────────────────────────────────────────────────────
@app.route('/api/merge-orders', methods=['POST'])
def merge_orders():
    """
    Upload a new order file and add its stops to the existing plan.

    Logic:
      1. Parse new file with the same pipeline as Generate Route Plan.
      2. Build a set of address+WH keys already present in the plan.
      3. Skip stops already covered; collect truly new stops.
      4. Run Clarke-Wright on new stops grouped by wh_group.
      5. For each new route try to fit into an existing truck (best-fit
         bin-packing: pick the truck with the most remaining capacity
         that still fits all new stops).
      6. If no truck fits, create a new truck.
      7. Re-number trucks, save to temp plan, return updated plan.
    """
    import time as _mtime

    router    = request.args.get('router', 'default')
    temp_plan = get_router_temp('plan', router)

    if not os.path.exists(temp_plan):
        return jsonify({'error': 'No existing plan found — generate a base plan first'}), 400

    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    fname = secure_filename(file.filename)
    upload_dir = app.config['UPLOAD_FOLDER']
    os.makedirs(upload_dir, exist_ok=True)
    fpath = os.path.join(upload_dir, f'merge_{fname}')
    file.save(fpath)

    try:
        # ── Load existing plan ─────────────────────────────
        with open(temp_plan) as fp:
            existing_trucks = json.load(fp)

        # Build lookup of SO numbers already in plan (dedup by SO, not by address)
        existing_so_nums = set()
        for t in existing_trucks:
            for s in t.get('stops', []):
                dn = (s.get('doc_number') or '').upper().strip()
                if dn: existing_so_nums.add(dn)
                for r in s.get('rows', []):
                    pdn = (r.get('_pfc_doc_num') or '').upper().strip()
                    if pdn: existing_so_nums.add(pdn)
        # Fallback addr+wh dedup for rows with no SO number
        existing_addr_keys = set()
        for t in existing_trucks:
            for s in t.get('stops', []):
                if not (s.get('doc_number') or '').strip():
                    ak = re.sub(r'\s+', ' ',
                        (s.get('shipping_address') or s.get('customer_name') or '').upper().strip())
                    existing_addr_keys.add(ak + '|||' + s.get('wh', ''))

        # ── Parse new file ─────────────────────────────────
        pd = _get_pd()
        try:
            xf = pd.ExcelFile(fpath)
            target_sheet = next(
                (s for s in xf.sheet_names
                 if 'assign' in s.lower() or 'order' in s.lower() or 'route' in s.lower()),
                xf.sheet_names[0])
            new_df = pd.read_excel(fpath, sheet_name=target_sheet, dtype=str)
        except Exception:
            try:
                new_df = pd.read_csv(fpath, dtype=str, encoding='utf-8')
            except Exception:
                new_df = pd.read_csv(fpath, dtype=str, encoding='latin1')
        finally:
            try: os.remove(fpath)
            except Exception: pass

        new_df.columns = [str(c).strip() for c in new_df.columns]
        col_map = normalize_headers(new_df.columns.tolist())

        # ── Build new stops dict (same logic as run_routing_engine) ──
        stops_dict = {}
        for _, row in new_df.iterrows():
            row_dict = row.to_dict()
            try:
                vol = float(safe_get(row_dict, col_map, 'tfor_qty', 0) or
                            safe_get(row_dict, col_map, 'do_qty', 0) or 0)
            except Exception:
                vol = 0
            if vol <= 0:
                continue

            cust    = str(safe_get(row_dict, col_map, 'customer_name', '')).strip() or 'UNKNOWN'
            addr    = str(safe_get(row_dict, col_map, 'shipping_address', '')).strip() or cust
            doc_num = str(safe_get(row_dict, col_map, 'so_number', '')).strip().upper()
            is_xfer = doc_num.startswith('TFORPFC')
            addr_key = re.sub(r'\s+', ' ', addr.upper().strip())
            loc_raw  = str(safe_get(row_dict, col_map, 'location', 'NA'))
            wh       = parse_location_cell(loc_raw)
            key      = addr_key + '|||' + wh

            # Skip if SO number already in plan; fallback to addr+wh for no-SO rows
            if doc_num and doc_num in existing_so_nums:
                continue
            if not doc_num and key in existing_addr_keys:
                continue

            cluster = str(safe_get(row_dict, col_map, 'cluster_id', 'UN')).strip().upper()
            area    = str(safe_get(row_dict, col_map, 'area', ''))
            try:
                lat = float(safe_get(row_dict, col_map, 'delivery_latitude', 0) or 0)
                lng = float(safe_get(row_dict, col_map, 'delivery_longitude', 0) or 0)
            except Exception:
                lat, lng = 0.0, 0.0
            s = serialise_row(row_dict)
            s['_pfc_doc_num'] = doc_num
            s['_pfc_vol'] = vol

            if key not in stops_dict:
                stops_dict[key] = {
                    'rows': [s], 'vol': vol, 'cluster_id': cluster,
                    'wh': wh, 'wh_group': get_wh_group(wh),
                    'match_area': clean_area_name(area),
                    'lat': lat, 'lng': lng,
                    'customer_name': cust, 'shipping_address': addr, 'area': area,
                    'all_customers': [cust],
                    'is_stock_transfer': is_xfer,
                    'doc_number': doc_num,
                }
            else:
                stops_dict[key]['vol'] += vol
                stops_dict[key]['rows'].append(s)
                if cust not in stops_dict[key]['all_customers']:
                    stops_dict[key]['all_customers'].append(cust)
                stops_dict[key]['customer_name'] = ' / '.join(stops_dict[key]['all_customers'])
                if is_xfer:
                    stops_dict[key]['is_stock_transfer'] = True

        new_stops = list(stops_dict.values())

        if not new_stops:
            return jsonify({
                'trucks': existing_trucks,
                'added': 0,
                'skipped': 0,
                'message': 'No new stops found — all addresses already exist in the current plan.'
            })

        # ── Geocode new stops (address → lat/lng) ─────────
        import time as _gtime2
        conn_geo = open_db()
        c_geo    = conn_geo.cursor()
        for stop in new_stops:
            addr2     = stop.get('shipping_address', '').strip()
            cache_key = re.sub(r'\s+', ' ', addr2.upper().strip()) if addr2 else ''

            # Priority 1: coords from the uploaded file
            if stop.get('lat') and stop.get('lng'):
                if cache_key:
                    c_geo.execute(
                        'INSERT OR IGNORE INTO address_fwd_cache (address_key,lat,lng,display_name) VALUES (?,?,?,?)',
                        (cache_key, stop['lat'], stop['lng'], addr2))
                    conn_geo.commit()
                continue  # no geocoding needed

            if not addr2:
                continue

            # Priority 2: address cache
            c_geo.execute('SELECT lat, lng FROM address_fwd_cache WHERE address_key=?', (cache_key,))
            row = c_geo.fetchone()
            if row and row[0] and row[1]:
                stop['lat'] = float(row[0]); stop['lng'] = float(row[1])
                continue

            # Priority 3: Nominatim fallback
            for attempt in _address_candidates(addr2):
                try:
                    hit = _nominatim_search(attempt)
                    _gtime2.sleep(1.05)
                    if hit:
                        stop['lat'] = round(float(hit['lat']), 6)
                        stop['lng'] = round(float(hit['lon']), 6)
                        c_geo.execute(
                            'INSERT OR REPLACE INTO address_fwd_cache (address_key,lat,lng,display_name) VALUES (?,?,?,?)',
                            (cache_key, stop['lat'], stop['lng'], hit.get('display_name', '')))
                        conn_geo.commit()
                        break
                except Exception:
                    _gtime2.sleep(1.05)
        conn_geo.close()

        # ── Route new stops by wh_group ────────────────────
        by_wh_group = defaultdict(list)
        for stop in new_stops:
            by_wh_group[stop['wh_group']].append(stop)

        added_stops = 0
        new_trucks  = []

        # Current truck index (for numbering new trucks)
        max_idx = 0
        for t in existing_trucks:
            try:
                n = int(t['truck_id'].replace('TRK-', ''))
                max_idx = max(max_idx, n)
            except Exception:
                pass

        for wh_group, group_stops in by_wh_group.items():
            wh_code = next(
                (code for code, d in WAREHOUSE_DATA.items() if d['group'] == wh_group),
                'FCSC')
            wh_lat, wh_lng = get_wh_coords(wh_code)

            STD_CAP   = 5000
            std_stops = [s for s in group_stops if s['vol'] <= STD_CAP * SETTINGS['cap_tolerance']]
            heavy     = [s for s in group_stops if s['vol']  > STD_CAP * SETTINGS['cap_tolerance']]

            routes = clarke_wright_routing(wh_lat, wh_lng, std_stops, target_cap=STD_CAP)
            for hs in heavy:
                routes.append([hs])

            # Eligible existing trucks for this wh_group (skip locked trucks)
            existing_wh_trucks = [
                t for t in existing_trucks
                if t.get('wh_group') == wh_group and not t.get('locked', False)
            ]

            for new_route_stops in routes:
                route_vol = sum(s['vol'] for s in new_route_stops)

                # Best-fit: find existing truck with most spare capacity that fits
                best_truck = None
                best_spare = -1
                for et in existing_wh_trucks:
                    cap     = et.get('truck_cap', 5000)
                    loaded  = et.get('acc_weight', 0)
                    spare   = cap * SETTINGS['cap_tolerance'] - loaded
                    n_drops = len(et.get('stops', []))
                    vt      = best_truck_type(loaded + route_vol, n_drops + len(new_route_stops))
                    if (spare >= route_vol and
                            n_drops + len(new_route_stops) <= vt['max_drops'] and
                            spare > best_spare):
                        best_truck = et
                        best_spare = spare

                if best_truck:
                    # Merge into existing truck
                    cur_seq = len(best_truck['stops'])
                    new_vol = best_truck['acc_weight'] + route_vol
                    new_cap = best_truck_type(new_vol, len(best_truck['stops']) + len(new_route_stops))
                    best_truck['acc_weight'] = round(new_vol, 3)
                    best_truck['truck_type'] = new_cap['label']
                    best_truck['truck_cap']  = new_cap['cap']
                    best_truck['rated_cap']  = new_cap['cap']
                    best_truck['util_pct']   = round(new_vol / new_cap['cap'], 4)

                    for si, stop in enumerate(new_route_stops):
                        cur_seq += 1
                        seq = cur_seq
                        add_fee = 0
                        if new_cap['label'] == '5MT':
                            add_fee = SETTINGS['add_drop_fee'] if seq > SETTINGS['add_drop_threshold_5mt'] else 0
                        elif new_cap['label'] == '2.5MT':
                            add_fee = SETTINGS['add_drop_fee'] if seq > SETTINGS['add_drop_threshold_25mt'] else 0
                        best_truck['stops'].append({
                            'seq': seq,
                            'cluster_id':    stop['cluster_id'],
                            'customer_name': stop['customer_name'],
                            'shipping_address': stop['shipping_address'],
                            'area':          stop['area'],
                            'match_area':    stop['match_area'],
                            'lat':           stop['lat'],
                            'lng':           stop['lng'],
                            'vol':           round(stop['vol'], 3),
                            'wh':            stop['wh'],
                            'add_drop_fee':  add_fee,
                            'is_stock_transfer': stop.get('is_stock_transfer', False),
                            'doc_number':    stop.get('doc_number', ''),
                            'all_customers': stop.get('all_customers', [stop['customer_name']]),
                            'rows':          stop['rows'],
                        })
                        added_stops += 1

                    # Re-run TSP on merged stops
                    reordered = nearest_neighbor_tsp(wh_lat, wh_lng, [
                        type('S', (), s)() if False else s
                        for s in best_truck['stops']
                    ])
                    for i, s in enumerate(reordered):
                        s['seq'] = i + 1
                    best_truck['stops'] = reordered
                    best_truck['total_km'] = calculate_total_km(wh_lat, wh_lng, best_truck['stops'])

                else:
                    # Create a new truck
                    max_idx += 1
                    truck_vol = route_vol
                    v_type    = best_truck_type(truck_vol, len(new_route_stops))
                    ordered   = nearest_neighbor_tsp(wh_lat, wh_lng, new_route_stops)
                    total_km  = calculate_total_km(wh_lat, wh_lng, ordered)

                    wh_vols = {}
                    for s in ordered:
                        wh_vols[s['wh']] = wh_vols.get(s['wh'], 0) + s['vol']
                    primary_wh = max(wh_vols, key=wh_vols.get) if wh_vols else wh_code

                    stops_out = []
                    for si, stop in enumerate(ordered):
                        seq = si + 1
                        add_fee = 0
                        if v_type['label'] == '5MT':
                            add_fee = SETTINGS['add_drop_fee'] if seq > SETTINGS['add_drop_threshold_5mt'] else 0
                        elif v_type['label'] == '2.5MT':
                            add_fee = SETTINGS['add_drop_fee'] if seq > SETTINGS['add_drop_threshold_25mt'] else 0
                        stops_out.append({
                            'seq': seq,
                            'cluster_id':    stop['cluster_id'],
                            'customer_name': stop['customer_name'],
                            'shipping_address': stop['shipping_address'],
                            'area':          stop['area'],
                            'match_area':    stop['match_area'],
                            'lat':           stop['lat'],
                            'lng':           stop['lng'],
                            'vol':           round(stop['vol'], 3),
                            'wh':            stop['wh'],
                            'add_drop_fee':  add_fee,
                            'is_stock_transfer': stop.get('is_stock_transfer', False),
                            'doc_number':    stop.get('doc_number', ''),
                            'all_customers': stop.get('all_customers', [stop['customer_name']]),
                            'rows':          stop['rows'],
                        })
                        added_stops += 1

                    new_trucks.append({
                        'truck_id':    f'TRK-{max_idx:03d}',
                        'truck_type':  v_type['label'],
                        'truck_cap':   v_type['cap'],
                        'rated_cap':   v_type['cap'],
                        'acc_weight':  round(truck_vol, 3),
                        'util_pct':    round(truck_vol / v_type['cap'], 4),
                        'pickup_wh':   primary_wh,
                        'all_whs':     sorted(wh_vols.keys()),
                        'wh_group':    wh_group,
                        'trucker_code': '',
                        'ref_number':   '',
                        'total_km':    total_km,
                        'stops':       stops_out,
                    })

        # Merge new trucks into plan and renumber
        updated_plan = existing_trucks + new_trucks
        for i, t in enumerate(updated_plan):
            t['truck_id'] = f'TRK-{i+1:03d}'

        with open(temp_plan, 'w') as fp:
            json.dump(updated_plan, fp, default=str)
        # ── Broadcast merged plan to all browsers ──
        merge_router = request.args.get('router', 'default')
        merge_date   = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
        set_working_plan(merge_date, updated_plan, merge_router)

        return jsonify({
            'trucks':     updated_plan,
            'added':      added_stops,
            'new_trucks': len(new_trucks),
            'message':    (f'Added {added_stops} new stop(s) across {len(new_trucks)} new truck(s). '
                           f'Stops already in the plan were skipped.'),
        })

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


# ─────────────────────────────────────────────────────────────
#  CSV EXPORTS
# ─────────────────────────────────────────────────────────────
@app.route('/api/export-csv', methods=['GET'])
def export_csv():
    import io as _io, csv as _csv
    router    = request.args.get('router', 'default')
    temp_plan = get_router_temp('plan', router)
    if not os.path.exists(temp_plan):
        return jsonify({'error': 'No plan loaded'}), 400
    with open(temp_plan) as fp:
        trucks = json.load(fp)
    buf = _io.StringIO()
    w   = _csv.writer(buf)
    w.writerow(['Truck ID','Truck Type','Truck Cap (kg)','Acc Weight (kg)','Util %',
                'Pickup WH','Trucker Code','Ref Number','Total KM',
                'Seq','Cluster ID','Customer Name','Shipping Address','Area',
                'Volume (kg)','WH','Add Drop Fee','Is Stock Transfer','Doc Number',
                'Latitude','Longitude'])
    for t in trucks:
        for s in t.get('stops', []):
            w.writerow([
                t.get('truck_id',''), t.get('truck_type',''), t.get('truck_cap',''),
                round(t.get('acc_weight',0),3), round((t.get('util_pct') or 0)*100,1),
                t.get('pickup_wh',''), t.get('trucker_code',''), t.get('ref_number',''),
                t.get('total_km',''), s.get('seq',''), s.get('cluster_id',''),
                s.get('customer_name',''), s.get('shipping_address',''), s.get('area',''),
                round(s.get('vol',0),3), s.get('wh',''), s.get('add_drop_fee',0),
                'YES' if s.get('is_stock_transfer') else '', s.get('doc_number',''),
                s.get('lat',''), s.get('lng',''),
            ])
    date_str = datetime.now().strftime('%Y-%m-%d')
    output   = _io.BytesIO(buf.getvalue().encode('utf-8-sig'))
    return send_file(output, mimetype='text/csv',
                     as_attachment=True, download_name=f'PFC_Route_Plan_{date_str}.csv')


@app.route('/api/export-monitoring-csv', methods=['GET'])
def export_monitoring_csv():
    import io as _io, csv as _csv
    router   = request.args.get('router', 'default')
    temp_mon = get_router_temp('monitoring', router)
    rows     = []
    if os.path.exists(temp_mon):
        with open(temp_mon) as fp:
            rows = json.load(fp)
    buf = _io.StringIO()
    w   = _csv.writer(buf)
    w.writerow(['Truck ID','Truck Type','Seq','Customer Name','Cluster ID','Area',
                'Status','Receiving Time','Actual Done','OTIF Status','Concerns','Remarks',
                'Trucker Code','Plan Date'])
    for r in rows:
        w.writerow([
            r.get('truck_id',''), r.get('truck_type',''), r.get('seq',''),
            r.get('customer_name',''), r.get('cluster_id',''), r.get('area',''),
            r.get('status',''), r.get('receiving_time',''), r.get('actual_done',''),
            r.get('otif_status',''), r.get('concerns',''), r.get('remarks',''),
            r.get('trucker_code',''), r.get('plan_date',''),
        ])
    date_str = datetime.now().strftime('%Y-%m-%d')
    output   = _io.BytesIO(buf.getvalue().encode('utf-8-sig'))
    return send_file(output, mimetype='text/csv',
                     as_attachment=True, download_name=f'PFC_Monitoring_{date_str}.csv')


@app.route('/api/clear-coord-cache', methods=['POST'])
def clear_coord_cache():
    """Wipe the forward-geocoding cache (address -> lat/lng)."""
    conn = open_db(); c = conn.cursor()
    c.execute('DELETE FROM address_fwd_cache')
    deleted = c.rowcount
    conn.commit(); conn.close()
    return jsonify({'deleted': deleted, 'message': f'Cleared {deleted} cached address(es).'})


@app.route('/api/verify-addresses', methods=['POST'])
def verify_addresses():
    """Reverse-geocode each stop's lat/lng via Nominatim and compare with stated area."""
    import urllib.request, urllib.parse, time as _time
    data  = request.get_json() or {}
    stops = data.get('stops', [])
    if not stops:
        return jsonify({'results': []})
    conn = open_db(); c = conn.cursor()
    results = []
    for stop in stops:
        lat = stop.get('lat'); lng = stop.get('lng')
        stated_raw = stop.get('area', '')
        if not lat or not lng:
            results.append({**stop, 'geocoded_city': '', 'geocoded_display': '',
                            'match': None, 'note': 'no coordinates'})
            continue
        lat_r = round(float(lat), 5); lng_r = round(float(lng), 5)
        c.execute('SELECT city, municipality, province, display_name FROM geocode_cache WHERE lat=? AND lng=?',
                  (lat_r, lng_r))
        cached = c.fetchone()
        if cached:
            city, muni, prov, display = cached
        else:
            try:
                url = (f'https://nominatim.openstreetmap.org/reverse'
                       f'?lat={lat_r}&lon={lng_r}&format=json&addressdetails=1')
                req = urllib.request.Request(url, headers={'User-Agent': 'PFC-Routing/1.0 supplychain3@premierfoodchoice.com'})
                with urllib.request.urlopen(req, timeout=6) as resp:
                    geo = json.loads(resp.read())
                addr    = geo.get('address', {})
                city    = (addr.get('city') or addr.get('municipality') or
                           addr.get('town') or addr.get('village') or addr.get('suburb') or '')
                muni    = addr.get('municipality', '') or addr.get('city_district', '')
                prov    = addr.get('state', '')
                display = geo.get('display_name', '')
                c.execute('INSERT OR REPLACE INTO geocode_cache (lat,lng,city,municipality,province,display_name) VALUES (?,?,?,?,?,?)',
                          (lat_r, lng_r, city, muni, prov, display))
                conn.commit()
                _time.sleep(1.05)
            except Exception as e:
                results.append({**stop, 'geocoded_city': '', 'geocoded_display': '',
                                'match': None, 'note': f'geocode error: {e}'})
                continue
        stated_candidates = [c2.upper() for c2 in area_candidates(stated_raw)]
        geo_tokens = [t.upper().strip() for t in
                      [city, muni, prov] + (city+' '+muni).split() if t.strip()]
        match = any(sc in geo_tokens or any(sc in gt or gt in sc for gt in geo_tokens)
                    for sc in stated_candidates if sc)
        geocoded_city = city or muni or prov or '?'
        results.append({
            **stop,
            'geocoded_city':    geocoded_city,
            'geocoded_display': display,
            'match':            match,
            'note':             '' if match else f'Stated "{stated_raw}" != geocoded "{geocoded_city}"',
        })
    conn.close()
    return jsonify({'results': results})


@app.route('/api/geocode-search', methods=['GET'])
def geocode_search():
    """Forward geocode a free-text query via Nominatim — returns top 5 candidates."""
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify({'results': []})
    try:
        url = (f'https://nominatim.openstreetmap.org/search'
               f'?q={urllib.parse.quote(q)}&format=json&addressdetails=1'
               f'&countrycodes=ph&limit=5')
        req = urllib.request.Request(
            url, headers={'User-Agent': 'PFC-Routing/1.0 supplychain3@premierfoodchoice.com'})
        with urllib.request.urlopen(req, timeout=8) as resp:
            hits = json.loads(resp.read())
        results = []
        for h in hits:
            results.append({
                'display': h.get('display_name', ''),
                'lat': float(h.get('lat', 0)),
                'lng': float(h.get('lon', 0)),
            })
        return jsonify({'results': results})
    except Exception as e:
        return jsonify({'results': [], 'error': str(e)})


@app.route('/api/update-stop-coord', methods=['POST'])
def update_stop_coord():
    """Patch lat/lng for one stop in the router's temp plan and persist to DB cache."""
    data   = request.get_json() or {}
    router = data.get('router', 'default')
    ti     = data.get('truck_index')
    si     = data.get('stop_index')
    new_lat = data.get('lat')
    new_lng = data.get('lng')
    if ti is None or si is None or new_lat is None or new_lng is None:
        return jsonify({'error': 'Missing fields'}), 400

    temp_plan = get_router_temp('plan', router)
    if not os.path.exists(temp_plan):
        return jsonify({'error': 'No active plan'}), 404

    with open(temp_plan) as f:
        trucks = json.load(f)
    new_name = data.get('customer_name')
    new_addr = data.get('shipping_address')
    try:
        stop = trucks[ti]['stops'][si]
        stop['lat'] = round(float(new_lat), 6)
        stop['lng'] = round(float(new_lng), 6)
        if new_name:
            stop['customer_name'] = new_name
        if new_addr:
            stop['shipping_address'] = new_addr
        # Cache corrected coords under the (new) address key
        addr = stop.get('shipping_address') or stop.get('customer_name') or ''
        if addr:
            cache_key = re.sub(r'\s+', ' ', addr.upper().strip())
            conn = open_db(); c = conn.cursor()
            c.execute('INSERT OR REPLACE INTO address_fwd_cache (address_key,lat,lng,display_name) VALUES (?,?,?,?)',
                      (cache_key, stop['lat'], stop['lng'], addr))
            conn.commit(); conn.close()
        with open(temp_plan, 'w') as f:
            json.dump(trucks, f, default=str)
        return jsonify({'ok': True, 'stop': stop})
    except (IndexError, KeyError) as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/fix-coords', methods=['POST'])
def fix_coords():
    """Re-geocode stops from shipping address via Nominatim."""
    import time as _time
    data  = request.get_json() or {}
    stops = data.get('stops', [])
    if not stops:
        return jsonify({'results': [], 'fixed': 0, 'total': 0})
    conn = open_db(); c = conn.cursor()
    results = []
    for stop in stops:
        addr = stop.get('shipping_address') or stop.get('customer_name') or ''
        if not addr:
            results.append({**stop, 'new_lat': None, 'new_lng': None, 'display_name': '', 'status': 'no address'})
            continue
        cache_key = re.sub(r'\s+', ' ', addr.upper().strip())
        c.execute('SELECT lat, lng, display_name FROM address_fwd_cache WHERE address_key=?', (cache_key,))
        row = c.fetchone()
        if row and row[0] and row[1]:
            results.append({**stop, 'new_lat': float(row[0]), 'new_lng': float(row[1]),
                            'display_name': row[2] or '', 'status': 'cached'})
            continue
        new_lat = new_lng = None; display = ''; status = 'not found'
        for attempt in _address_candidates(addr):
            try:
                hit = _nominatim_search(attempt)
                _time.sleep(1.05)
                if hit:
                    new_lat = round(float(hit['lat']), 6)
                    new_lng = round(float(hit['lon']), 6)
                    display = hit.get('display_name', '')
                    status  = 'geocoded'
                    c.execute('INSERT OR REPLACE INTO address_fwd_cache (address_key,lat,lng,display_name) VALUES (?,?,?,?)',
                              (cache_key, new_lat, new_lng, display))
                    conn.commit()
                    break
                else:
                    _time.sleep(1.05)
            except Exception as e:
                results.append({**stop, 'new_lat': None, 'new_lng': None,
                                'display_name': '', 'status': f'error: {e}'})
                new_lat = None; break
        results.append({**stop, 'new_lat': new_lat, 'new_lng': new_lng,
                        'display_name': display, 'status': status})
    conn.close()
    fixed = sum(1 for r in results if r.get('new_lat') is not None)
    return jsonify({'results': results, 'fixed': fixed, 'total': len(stops)})




# ── Real-time collaboration endpoints ─────────────────────────────────────────
@app.route('/api/collab/push', methods=['POST'])
def collab_push():
    data = request.get_json() or {}
    cid  = (data.get('collab_id') or '').strip()
    if not cid:
        return jsonify({'error': 'missing collab_id'}), 400
    _collab[cid] = {
        'trucks':  data.get('trucks', []),
        'version': time.time(),
        'by':      data.get('by', ''),
    }
    return jsonify({'success': True, 'version': _collab[cid]['version']})


@app.route('/api/collab/poll', methods=['GET'])
def collab_poll():
    cid   = (request.args.get('collab_id') or '').strip()
    since = float(request.args.get('since', 0) or 0)
    if cid not in _collab:
        return jsonify({'updated': False})
    state = _collab[cid]
    if state['version'] <= since:
        return jsonify({'updated': False, 'version': state['version'], 'by': state['by']})
    return jsonify({
        'updated': True,
        'version': state['version'],
        'by':      state['by'],
        'trucks':  state['trucks'],
    })


@app.route('/api/collab/heartbeat', methods=['POST'])
def collab_heartbeat():
    """Track who is actively viewing a collab session."""
    data = request.get_json() or {}
    cid  = (data.get('collab_id') or '').strip()
    name = (data.get('by') or '').strip()
    if not cid or not name:
        return jsonify({'error': 'missing fields'}), 400
    key = cid + '__presence'
    if key not in _collab:
        _collab[key] = {}
    _collab[key][name] = time.time()
    # Expire users not seen in 15 s
    cutoff = time.time() - 15
    _collab[key] = {k: v for k, v in _collab[key].items() if v > cutoff}
    return jsonify({'users': list(_collab[key].keys())})


# GOOGLE SHEETS SYNC
GSHEETS_CREDS_PATH = os.path.join(os.path.dirname(__file__), 'google_credentials.json')
GSHEETS_SCOPES     = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive',
]

# ── OAuth-based Google credential helpers ──────────────────────────────────────

GOOGLE_TOKEN_PATH = os.path.join(os.path.dirname(__file__), 'google_token.json')

def _get_google_creds():
    # Returns valid OAuth credentials, or raises RuntimeError('NOT_AUTHORIZED')
    try:
        from google.oauth2.credentials import Credentials
        from google.auth.transport.requests import Request
    except ImportError:
        raise RuntimeError(
            "Google API libraries not installed. "
            "Run: pip install google-api-python-client google-auth google-auth-httplib2 google-auth-oauthlib"
        )
    if not os.path.exists(GOOGLE_TOKEN_PATH):
        raise RuntimeError('NOT_AUTHORIZED')
    creds = Credentials.from_authorized_user_file(GOOGLE_TOKEN_PATH, GSHEETS_SCOPES)
    if creds.expired and creds.refresh_token:
        try:
            creds.refresh(Request())
            with open(GOOGLE_TOKEN_PATH, 'w') as f:
                f.write(creds.to_json())
        except Exception:
            raise RuntimeError('NOT_AUTHORIZED')
    if not creds.valid:
        raise RuntimeError('NOT_AUTHORIZED')
    return creds


def _get_gsheets_service():
    from googleapiclient.discovery import build
    return build('sheets', 'v4', credentials=_get_google_creds())


def _get_gdrive_service():
    from googleapiclient.discovery import build
    return build('drive', 'v3', credentials=_get_google_creds())


@app.route('/api/google/status', methods=['GET'])
def google_status():
    creds_file_ok = os.path.exists(GSHEETS_CREDS_PATH)
    authorized = False
    if os.path.exists(GOOGLE_TOKEN_PATH):
        try:
            from google.oauth2.credentials import Credentials
            from google.auth.transport.requests import Request
            creds = Credentials.from_authorized_user_file(GOOGLE_TOKEN_PATH, GSHEETS_SCOPES)
            if creds.expired and creds.refresh_token:
                creds.refresh(Request())
                with open(GOOGLE_TOKEN_PATH, 'w') as f:
                    f.write(creds.to_json())
            authorized = creds.valid
        except Exception:
            pass
    return jsonify({'creds_file_ok': creds_file_ok, 'authorized': authorized})


@app.route('/api/google/auth', methods=['GET'])
def google_auth_start():
    if not os.path.exists(GSHEETS_CREDS_PATH):
        return jsonify({'error': 'google_credentials.json not found'}), 400
    try:
        from google_auth_oauthlib.flow import Flow
    except ImportError:
        return jsonify({'error': 'google-auth-oauthlib not installed. Run: pip install google-auth-oauthlib'}), 500
    port = int(os.environ.get('PORT', 5050))
    redirect_uri = 'http://localhost:{}/api/google/callback'.format(port)
    flow = Flow.from_client_secrets_file(
        GSHEETS_CREDS_PATH, scopes=GSHEETS_SCOPES, redirect_uri=redirect_uri
    )
    auth_url, state = flow.authorization_url(
        access_type='offline', include_granted_scopes='true', prompt='consent'
    )
    state_path = os.path.join(os.path.dirname(__file__), 'google_state.tmp')
    with open(state_path, 'w') as f:
        f.write(state)
    from flask import redirect as flask_redirect
    return flask_redirect(auth_url)


@app.route('/api/google/callback', methods=['GET'])
def google_auth_callback():
    if not os.path.exists(GSHEETS_CREDS_PATH):
        return 'google_credentials.json not found', 400
    try:
        from google_auth_oauthlib.flow import Flow
    except ImportError:
        return 'google-auth-oauthlib not installed', 500
    state_path = os.path.join(os.path.dirname(__file__), 'google_state.tmp')
    state = ''
    if os.path.exists(state_path):
        with open(state_path) as f:
            state = f.read().strip()
    port = int(os.environ.get('PORT', 5050))
    redirect_uri = 'http://localhost:{}/api/google/callback'.format(port)
    flow = Flow.from_client_secrets_file(
        GSHEETS_CREDS_PATH, scopes=GSHEETS_SCOPES, state=state, redirect_uri=redirect_uri
    )
    import os as _os
    _os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '1'
    flow.fetch_token(authorization_response=request.url.replace('https://', 'http://'))
    creds = flow.credentials
    with open(GOOGLE_TOKEN_PATH, 'w') as f:
        f.write(creds.to_json())
    return (
        '<html><body style="font-family:sans-serif;text-align:center;padding:60px">'
        '<h2 style="color:#16a34a">&#10003; Google Account Connected!</h2>'
        '<p>You can close this tab and return to the app.</p>'
        '<script>setTimeout(function(){window.close();},2000);</script>'
        '</body></html>'
    )


@app.route('/api/google/disconnect', methods=['POST'])
def google_disconnect():
    if os.path.exists(GOOGLE_TOKEN_PATH):
        os.remove(GOOGLE_TOKEN_PATH)
    return jsonify({'success': True})



def _gdrive_get_or_create_folder(svc, name, parent_id=None):
    # Return folder ID for name inside parent_id, creating if needed
    q = "mimeType='application/vnd.google-apps.folder' and name='{}' and trashed=false".format(name)
    if parent_id:
        q += " and '{}' in parents".format(parent_id)
    res = svc.files().list(q=q, fields='files(id,name)', spaces='drive').execute()
    files = res.get('files', [])
    if files:
        return files[0]['id']
    meta = {'name': name, 'mimeType': 'application/vnd.google-apps.folder'}
    if parent_id:
        meta['parents'] = [parent_id]
    folder = svc.files().create(body=meta, fields='id').execute()
    return folder['id']


def _gdrive_build_excel(trucks, plan_date):
    # Build an in-memory Excel workbook from the truck plan and return bytes
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = 'Route Plan'
    headers = ['Date', 'Truck ID', 'Truck Type', 'WH', 'Seq',
               'Customer', 'Address', 'Area', 'Cluster',
               'Volume (KG)', 'SO Number', 'Trucker', 'Rate (PHP)', 'Stock Transfer']
    hdr_fill = PatternFill('solid', fgColor='1F4E79')
    hdr_font = Font(color='FFFFFF', bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')
    for truck in trucks:
        for stop in truck.get('stops', []):
            ws.append([
                plan_date,
                truck.get('truck_id', ''),
                truck.get('truck_type', ''),
                stop.get('wh', ''),
                stop.get('seq', ''),
                stop.get('customer_name', ''),
                stop.get('shipping_address', ''),
                stop.get('match_area', '') or stop.get('area', ''),
                stop.get('cluster_id', ''),
                round(float(stop.get('vol', 0) or 0), 2),
                stop.get('doc_number', ''),
                truck.get('trucker_code', ''),
                truck.get('truck_rate', ''),
                'YES' if stop.get('is_stock_transfer') else '',
            ])
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _gdrive_upload_plan_excel(trucks, plan_date, plan_name):
    # Upload plan Excel to Drive under PFC Route Plans/YYYY-MM-DD/. Returns webViewLink.
    import io
    try:
        from googleapiclient.http import MediaIoBaseUpload
    except ImportError:
        return None
    cfg_path = os.path.join(os.path.dirname(__file__), 'gdrive_config.json')
    root_folder_id = None
    if os.path.exists(cfg_path):
        try:
            with open(cfg_path) as fp:
                root_folder_id = json.load(fp).get('folder_id') or None
        except Exception:
            pass
    svc = _get_gdrive_service()
    root_id = _gdrive_get_or_create_folder(svc, 'PFC Route Plans', root_folder_id)
    date_id = _gdrive_get_or_create_folder(svc, plan_date, root_id)
    xlsx_bytes = _gdrive_build_excel(trucks, plan_date)
    safe_name = plan_name.replace('/', '-').replace(':', '-')
    file_name = safe_name + '.xlsx'
    media = MediaIoBaseUpload(
        io.BytesIO(xlsx_bytes),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        resumable=False
    )
    file_meta = {'name': file_name, 'parents': [date_id]}
    created = svc.files().create(body=file_meta, media_body=media, fields='id,webViewLink').execute()
    try:
        svc.permissions().create(
            fileId=created['id'],
            body={'type': 'anyone', 'role': 'reader'}
        ).execute()
    except Exception:
        pass
    return created.get('webViewLink')


@app.route('/api/gdrive/config', methods=['GET'])
def gdrive_get_config():
    creds_ok = os.path.exists(GOOGLE_TOKEN_PATH)
    cfg_path = os.path.join(os.path.dirname(__file__), 'gdrive_config.json')
    folder_id = ''
    if os.path.exists(cfg_path):
        try:

            with open(cfg_path) as fp:
                folder_id = json.load(fp).get('folder_id') or ''
        except Exception:
            pass
    return jsonify({'creds_ok': creds_ok, 'folder_id': folder_id})

@app.route('/api/gdrive/config', methods=['POST'])
def gdrive_set_config():
    data = request.get_json(silent=True) or {}
    folder_id = (data.get('folder_id') or '').strip()
    cfg_path = os.path.join(os.path.dirname(__file__), 'gdrive_config.json')
    try:
        with open(cfg_path, 'w') as fp:
            json.dump({'folder_id': folder_id}, fp)
        return jsonify({'success': True})
    except Exception as ex:
        return jsonify({'success': False, 'error': str(ex)}), 500

# ── Bootstrap ───────────────────────────────────────────────────────────────
init_db()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
